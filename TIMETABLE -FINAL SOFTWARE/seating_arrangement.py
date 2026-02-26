"""Seating arrangement generator for exam classrooms."""
import pandas as pd
import random
import os
from config import INPUT_DIR, OUTPUT_DIR
from file_manager import FileManager
from excel_loader import ExcelLoader
from exam_scheduler import ExamScheduler
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

class SeatingArrangementGenerator:
    """Generates seating arrangements for exam classrooms."""
    
    def __init__(self, data_frames, schedule_generator):
        self.dfs = data_frames
        self.schedule_gen = schedule_generator
        self.exam_scheduler = ExamScheduler(data_frames, schedule_generator)
        self.exam_classrooms = self.exam_scheduler.exam_classrooms
        self.student_data = self._load_student_data()
        self.exam_schedule = self._get_exam_schedule()
        self.classroom_capacities = self._get_classroom_capacities()
        
        # Performance optimization: Cache semester courses and student-course mappings
        self._semester_courses_cache = {}  # Cache for semester courses
        self._student_courses_cache = {}  # Cache for student courses: (roll_no, semester) -> courses
        self._course_students_index = {}  # Reverse index: (semester, course) -> [students]
        self._precompute_student_courses()
    
    def _load_student_data(self):
        """Load student data from student_data.xlsx."""
        student_df = self.dfs.get('student')
        if student_df is None or student_df.empty:
            # Try alternative key names
            for key in self.dfs.keys():
                if 'student' in key.lower():
                    student_df = self.dfs[key]
                    break
        
        if student_df is None or student_df.empty:
            print("WARNING: No student data found")
            return pd.DataFrame()
        
        return student_df
    
    def _get_exam_schedule(self):
        """Get exam schedule to know which courses are on which day/session."""
        exam_schedule = {}
        
        # Get Pre-Mid and Post-Mid courses
        pre_mid_courses = self.exam_scheduler.get_all_pre_mid_courses()
        post_mid_courses = self.exam_scheduler.get_all_post_mid_courses()
        
        # Schedule exams to get day/session assignments
        if not pre_mid_courses.empty:
            mid_fn_df, mid_an_df = self.exam_scheduler.schedule_exams(pre_mid_courses, num_days=7)
            exam_days = ['Saturday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Monday']
            
            # Parse FN schedule
            for day in exam_days:
                if day in mid_fn_df.columns:
                    courses_str = mid_fn_df[day].iloc[0] if not mid_fn_df.empty else ''
                    if courses_str and str(courses_str).strip():
                        courses = [c.strip() for c in str(courses_str).split(',')]
                        for course in courses:
                            if course:
                                key = (day, 'FN', course.strip())
                                exam_schedule[key] = 'Pre-Mid'
            
            # Parse AN schedule
            for day in exam_days:
                if day in mid_an_df.columns:
                    courses_str = mid_an_df[day].iloc[0] if not mid_an_df.empty else ''
                    if courses_str and str(courses_str).strip():
                        courses = [c.strip() for c in str(courses_str).split(',')]
                        for course in courses:
                            if course:
                                key = (day, 'AN', course.strip())
                                exam_schedule[key] = 'Pre-Mid'
        
        if not post_mid_courses.empty:
            end_fn_df, end_an_df = self.exam_scheduler.schedule_exams(post_mid_courses, num_days=7)
            exam_days = ['Saturday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Monday']
            
            # Parse FN schedule
            for day in exam_days:
                if day in end_fn_df.columns:
                    courses_str = end_fn_df[day].iloc[0] if not end_fn_df.empty else ''
                    if courses_str and str(courses_str).strip():
                        courses = [c.strip() for c in str(courses_str).split(',')]
                        for course in courses:
                            if course:
                                key = (day, 'FN', course.strip())
                                exam_schedule[key] = 'Post-Mid'
            
            # Parse AN schedule
            for day in exam_days:
                if day in end_an_df.columns:
                    courses_str = end_an_df[day].iloc[0] if not end_an_df.empty else ''
                    if courses_str and str(courses_str).strip():
                        courses = [c.strip() for c in str(courses_str).split(',')]
                        for course in courses:
                            if course:
                                key = (day, 'AN', course.strip())
                                exam_schedule[key] = 'Post-Mid'
        
        return exam_schedule
    
    def _get_classroom_capacities(self):
        """Get exam capacity for each exam classroom."""
        capacities = {}
        classroom_df = self.dfs.get('classroom')
        
        if classroom_df is None or classroom_df.empty:
            # Use default capacity (6 rows × 4 columns × 2 students = 48)
            for room in self.exam_classrooms:
                capacities[room] = 48  # Default exam capacity
            return capacities
        
        # Find room number and exam capacity columns
        room_col = None
        exam_cap_col = None
        cap_col = None  # Fallback to regular capacity
        
        for col in classroom_df.columns:
            col_lower = str(col).lower()
            if room_col is None and any(k in col_lower for k in ['room', 'number', 'name']):
                room_col = col
            if exam_cap_col is None and 'exam' in col_lower and 'cap' in col_lower:
                exam_cap_col = col
            if cap_col is None and 'cap' in col_lower and 'exam' not in col_lower:
                cap_col = col
        
        if room_col is None:
            room_col = classroom_df.columns[0]
        
        for _, row in classroom_df.iterrows():
            room_name = str(row.get(room_col, '')).strip()
            if room_name in self.exam_classrooms:
                # Prefer exam capacity, fallback to regular capacity
                capacity = 48  # Default
                
                if exam_cap_col:
                    try:
                        exam_cap = row.get(exam_cap_col)
                        if pd.notna(exam_cap) and str(exam_cap).strip() and str(exam_cap).strip().lower() not in ['nil', 'none', '0', '']:
                            capacity = int(float(exam_cap))
                        elif exam_cap_col and capacity == 48 and cap_col:
                            # Fallback to regular capacity if exam capacity is 0/nil
                            try:
                                reg_cap = int(float(row.get(cap_col, 48)))
                                # Use half of regular capacity as exam capacity (typical)
                                capacity = max(48, reg_cap // 2)
                            except:
                                pass
                    except:
                        pass
                elif cap_col:
                    try:
                        reg_cap = int(float(row.get(cap_col, 48)))
                        capacity = max(48, reg_cap // 2)  # Use half of regular capacity
                    except:
                        pass
                
                capacities[room_name] = capacity
        
        # Set default for rooms not found (6 rows × 4 columns × 2 = 48)
        for room in self.exam_classrooms:
            if room not in capacities:
                capacities[room] = 48
        
        return capacities
    
    def _precompute_student_courses(self):
        """Pre-compute all student courses and create reverse index for fast lookup."""
        if self.student_data.empty:
            return
        
        print("  Pre-computing student-course mappings...")
        
        # Get all unique semesters
        semesters = self.student_data['Semester'].dropna().unique().tolist()
        
        # Cache semester courses
        for semester in semesters:
            if semester not in self._semester_courses_cache:
                sem_courses = ExcelLoader.get_semester_courses(self.dfs, int(semester))
                self._semester_courses_cache[int(semester)] = sem_courses
        
        # Pre-compute student courses and build reverse index
        for _, student in self.student_data.iterrows():
            roll_no = str(student.get('Roll No', '')).strip()
            semester = student.get('Semester', None)
            
            if not roll_no or pd.isna(semester):
                continue
            
            semester = int(semester)
            cache_key = (roll_no, semester)
            
            if cache_key not in self._student_courses_cache:
                # Extract department from roll number
                dept_codes = []
                if 'BCS' in roll_no:
                    dept_codes = ['CSE', 'CSE-A', 'CSE-B']
                elif 'BDS' in roll_no:
                    dept_codes = ['DSAI']
                elif 'BEC' in roll_no:
                    dept_codes = ['ECE']
                
                courses = []
                if dept_codes and semester in self._semester_courses_cache:
                    sem_courses = self._semester_courses_cache[semester]
                    if not sem_courses.empty and 'Department' in sem_courses.columns:
                        for dept_code in dept_codes:
                            dept_mask = sem_courses['Department'].astype(str) == dept_code
                            dept_courses = sem_courses[dept_mask]
                            if 'Course Code' in dept_courses.columns:
                                dept_course_list = dept_courses['Course Code'].dropna().unique().tolist()
                                courses.extend([str(c).strip() for c in dept_course_list])
                
                courses = list(set([c for c in courses if c]))
                self._student_courses_cache[cache_key] = courses
                
                # Build reverse index: course -> students
                for course in courses:
                    index_key = (semester, course)
                    if index_key not in self._course_students_index:
                        self._course_students_index[index_key] = []
                    self._course_students_index[index_key].append({
                        'Roll No': roll_no,
                        'Name': student.get('Name', ''),
                        'Semester': semester,
                        'Department': student.get('Department', ''),
                        'Course': course
                    })
        
        print(f"  Cached {len(self._student_courses_cache)} student-course mappings")
    
    def _get_student_courses(self, roll_no, semester):
        """Get courses for a student (from cache)."""
        cache_key = (roll_no, semester)
        return self._student_courses_cache.get(cache_key, [])
    
    def _get_students_for_exam(self, day, session):
        """Get all students who have exams on this day/session (optimized using reverse index)."""
        students_with_exams = []
        students_added = set()  # Track to avoid duplicates
        
        # Get courses scheduled for this day/session
        exam_courses = []
        for (exam_day, exam_session, course), exam_type in self.exam_schedule.items():
            if exam_day == day and exam_session == session:
                exam_courses.append(course)
        
        if not exam_courses:
            return students_with_exams
        
        # Use reverse index for fast lookup
        for course in exam_courses:
            # Check all semesters for this course
            for semester in self._semester_courses_cache.keys():
                index_key = (semester, course)
                if index_key in self._course_students_index:
                    for student in self._course_students_index[index_key]:
                        roll_no = student['Roll No']
                        if roll_no not in students_added:
                            students_with_exams.append(student)
                            students_added.add(roll_no)
        
        return students_with_exams
    
    def _can_sit_together(self, student1, student2, day, session):
        """Check if two students can sit together (no exam conflict)."""
        # Different semesters - always OK
        if student1['Semester'] != student2['Semester']:
            return True
        
        # Same semester - check if they have same exam
        course1 = student1.get('Course', '')
        course2 = student2.get('Course', '')
        
        # If no course info, be safe and say no
        if not course1 or not course2:
            return False
        
        # Same course exam - conflict!
        if course1 == course2:
            return False
        
        # Different courses, same semester - OK
        return True
    
    def _generate_seating_for_room_with_students(self, room_name, capacity, students):
        """Generate seating arrangement for a specific room with given students list."""
        if not students:
            return pd.DataFrame()
        
        # Shuffle students for random distribution
        random.shuffle(students)
        
        # Calculate number of benches based on 6 rows × 4 columns layout
        # Each bench has 2 students, so max capacity = 6 rows × 4 columns × 2 = 48 students
        max_seats = 6 * 4 * 2  # 48 seats (6 rows × 4 columns × 2 students per bench)
        actual_capacity = min(capacity, max_seats)
        num_benches = (actual_capacity // 2)
        if actual_capacity % 2 == 1:
            num_benches += 1  # One extra bench for odd capacity
        
        # Ensure we don't exceed 6 rows × 4 columns = 24 benches
        num_benches = min(num_benches, 6 * 4)
        
        # Create seating arrangement
        seating_data = []
        bench_num = 1
        
        # Pair students: prefer different semesters, then same semester different courses
        unassigned = students.copy()
        random.shuffle(unassigned)  # Randomize for better distribution
        
        while bench_num <= num_benches and unassigned:
            if len(unassigned) == 1:
                # Only one student left - assign alone
                student = unassigned[0]
                seating_data.append({
                    'Bench': bench_num,
                    'COL1': student['Roll No'],
                    'COL2': ''
                })
                unassigned.remove(student)
                bench_num += 1
                break
            
            student1 = unassigned[0]
            compatible_found = False
            
            # Strategy 1: Find student from different semester (fastest check)
            sem1 = student1['Semester']
            for student2 in unassigned[1:]:
                if student2['Semester'] != sem1:
                    # Different semester - always compatible
                    seating_data.append({
                        'Bench': bench_num,
                        'COL1': student1['Roll No'],
                        'COL2': student2['Roll No']
                    })
                    unassigned.remove(student1)
                    unassigned.remove(student2)
                    compatible_found = True
                    bench_num += 1
                    break
            
            if not compatible_found:
                # Strategy 2: Find student from same semester but different course
                course1 = student1.get('Course', '')
                for student2 in unassigned[1:]:
                    if student2['Semester'] == sem1:
                        course2 = student2.get('Course', '')
                        if course1 and course2 and course1 != course2:
                            # Same semester, different courses - compatible
                            seating_data.append({
                                'Bench': bench_num,
                                'COL1': student1['Roll No'],
                                'COL2': student2['Roll No']
                            })
                            unassigned.remove(student1)
                            unassigned.remove(student2)
                            compatible_found = True
                            bench_num += 1
                            break
            
            if not compatible_found:
                # No compatible pair found - assign alone
                seating_data.append({
                    'Bench': bench_num,
                    'COL1': student1['Roll No'],
                    'COL2': ''
                })
                unassigned.remove(student1)
                bench_num += 1
        
        if not seating_data:
            return pd.DataFrame()
        
        df = pd.DataFrame(seating_data)
        return df
    
    def _generate_seating_for_room(self, room_name, capacity, day, session):
        """Generate seating arrangement for a specific room, day, and session (legacy method)."""
        students = self._get_students_for_exam(day, session)
        return self._generate_seating_for_room_with_students(room_name, capacity, students)
    
    def _get_date_for_day(self, day, exam_days):
        """Get date for a given exam day."""
        from datetime import datetime, timedelta
        
        # Find next Saturday from today (or use a fixed start date)
        start_date = datetime(2025, 9, 20)  # Saturday, Sept 20, 2025
        
        day_to_date = {}
        current_date = start_date
        
        for i, exam_day in enumerate(exam_days):
            if i == 0:
                day_to_date[exam_day] = start_date
                current_date = start_date
            else:
                prev_day = exam_days[i-1]
                if prev_day == 'Saturday':
                    days_to_add = 2  # Skip Sunday
                elif prev_day == 'Friday':
                    days_to_add = 3  # Skip weekend
                elif prev_day == 'Monday' and i > 1:
                    days_to_add = 1
                else:
                    days_to_add = 1
                
                current_date += timedelta(days=days_to_add)
                day_to_date[exam_day] = current_date
        
        if day in day_to_date:
            return day_to_date[day].strftime('%d/%m/%Y')
        return day
    
    def _create_seating_section(self, day, session, seating_df, exam_days):
        """Create a seating section for one day/session combination."""
        if seating_df.empty:
            return []
        
        rows = []
        
        # Get formatted date
        date_str = self._get_date_for_day(day, exam_days)
        
        # Section header: Day, Date, Session
        header_row = [f'{day} - {date_str} - {session}'] + [''] * 8
        rows.append(header_row)
        rows.append([''] * 9)  # Empty row
        
        # WINDOW label
        window_row = ['WINDOW'] + [''] * 8
        rows.append(window_row)
        rows.append([''] * 9)  # Empty row
        
        # Column headers
        col_header = ['', 'COL1', '', 'COL2', '', 'COL3', '', 'COL4', '']
        rows.append(col_header)
        
        # Arrange benches in 6 rows × 4 columns format
        max_benches = len(seating_df)
        num_rows = 6
        benches_per_col = num_rows  # 6 rows per column
        
        # Create 6 rows for seating
        for row_idx in range(num_rows):
            row_data = ['']  # First column is empty
            
            # COL1
            bench1_idx = row_idx
            if bench1_idx < max_benches:
                bench1 = seating_df.iloc[bench1_idx]
                row_data.append(bench1['COL1'])
                row_data.append(bench1['COL2'] if bench1['COL2'] else '')
            else:
                row_data.append('')
                row_data.append('')
            
            # COL2
            bench2_idx = row_idx + benches_per_col
            if bench2_idx < max_benches:
                bench2 = seating_df.iloc[bench2_idx]
                row_data.append(bench2['COL1'])
                row_data.append(bench2['COL2'] if bench2['COL2'] else '')
            else:
                row_data.append('')
                row_data.append('')
            
            # COL3
            bench3_idx = row_idx + benches_per_col * 2
            if bench3_idx < max_benches:
                bench3 = seating_df.iloc[bench3_idx]
                row_data.append(bench3['COL1'])
                row_data.append(bench3['COL2'] if bench3['COL2'] else '')
            else:
                row_data.append('')
                row_data.append('')
            
            # COL4
            bench4_idx = row_idx + benches_per_col * 3
            if bench4_idx < max_benches:
                bench4 = seating_df.iloc[bench4_idx]
                row_data.append(bench4['COL1'])
                row_data.append(bench4['COL2'] if bench4['COL2'] else '')
            else:
                row_data.append('')
                row_data.append('')
            
            rows.append(row_data)
        
        # Door label
        rows.append([''] * 9)  # Empty row
        door_row = ['Door'] + [''] * 8
        rows.append(door_row)
        rows.append([''] * 9)  # Empty row separator
        
        return rows
    
    def _create_seating_sheet(self, room_name, seating_data_by_day_session, exam_days):
        """Create a combined seating arrangement sheet for one room showing all days and sessions."""
        rows = []
        
        # Room header
        room_header = [f'Room: {room_name}'] + [''] * 8
        rows.append(room_header)
        rows.append([''] * 9)  # Empty row
        
        # Create sections for each day/session combination
        exam_days_list = ['Saturday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Monday']
        sessions = ['FN', 'AN']
        
        for day in exam_days_list:
            for session in sessions:
                key = (day, session)
                if key in seating_data_by_day_session:
                    seating_df = seating_data_by_day_session[key]
                    if not seating_df.empty:
                        section_rows = self._create_seating_section(day, session, seating_df, exam_days)
                        rows.extend(section_rows)
        
        # Create DataFrame
        max_cols = max(len(row) for row in rows) if rows else 9
        padded_rows = []
        for row in rows:
            padded_row = row + [''] * (max_cols - len(row))
            padded_rows.append(padded_row)
        
        df = pd.DataFrame(padded_rows)
        return df
    
    def _format_seating_sheet(self, worksheet, room_name, day, session):
        """Apply color coding and formatting to seating arrangement sheet."""
        try:
            from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
            
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            if max_row == 0 or max_col == 0:
                return
            
            # Define colors
            room_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue
            room_header_font = Font(bold=True, size=14, color="FFFFFF")  # White text
            
            section_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Dark blue
            section_header_font = Font(bold=True, size=11, color="FFFFFF")  # White text
            
            window_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")  # Light blue
            window_font = Font(bold=True, size=12, color="000000")
            
            door_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Very light blue
            door_font = Font(bold=True, size=11, color="000000")
            
            col1_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green
            col2_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow
            col3_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Light orange
            col4_fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")  # Light blue
            
            col_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Green
            col_header_font = Font(bold=True, size=10, color="FFFFFF")
            
            student_font = Font(size=9, color="000000")
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Find all section headers, window rows, door rows, and column headers
            room_header_rows = []
            section_header_rows = []
            window_rows = []
            door_rows = []
            col_header_rows = []
            
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell_value = str(cell.value).strip() if cell.value else ''
                    
                    if 'Room:' in cell_value and row_idx not in room_header_rows:
                        room_header_rows.append(row_idx)
                    if (('Saturday' in cell_value or 'Monday' in cell_value or 'Tuesday' in cell_value or 
                         'Wednesday' in cell_value or 'Thursday' in cell_value or 'Friday' in cell_value) and 
                        ('FN' in cell_value or 'AN' in cell_value)) and row_idx not in section_header_rows:
                        section_header_rows.append(row_idx)
                    if 'WINDOW' in cell_value.upper() and row_idx not in window_rows:
                        window_rows.append(row_idx)
                    if 'Door' in cell_value and row_idx not in door_rows:
                        door_rows.append(row_idx)
                    if ('COL1' in cell_value or 'COL2' in cell_value or 'COL3' in cell_value or 'COL4' in cell_value) and row_idx not in col_header_rows:
                        col_header_rows.append(row_idx)
            
            # Format room header rows
            for row_idx in room_header_rows:
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        cell.fill = room_header_fill
                        cell.font = room_header_font
                        cell.alignment = center_align
            
            # Format section header rows
            for row_idx in section_header_rows:
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        cell.fill = section_header_fill
                        cell.font = section_header_font
                        cell.alignment = center_align
            
            # Format WINDOW rows
            for row_idx in window_rows:
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = window_fill
                    cell.font = window_font
                    cell.alignment = center_align
            
            # Format Door rows
            for row_idx in door_rows:
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = door_fill
                    cell.font = door_font
                    cell.alignment = center_align
            
            # Format column headers
            for row_idx in col_header_rows:
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell_value = str(cell.value).strip() if cell.value else ''
                    if 'COL1' in cell_value or 'COL2' in cell_value or 'COL3' in cell_value or 'COL4' in cell_value:
                        cell.fill = col_header_fill
                        cell.font = col_header_font
                        cell.alignment = center_align
            
            # Format student data rows
            # Find data rows (between col headers and doors, excluding section headers)
            processed_rows = set(room_header_rows + section_header_rows + window_rows + door_rows + col_header_rows)
            
            for row_idx in range(1, max_row + 1):
                if row_idx in processed_rows:
                    continue
                
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell_value = str(cell.value).strip() if cell.value else ''
                    
                    # Skip empty cells
                    if not cell_value:
                        continue
                    
                    # Determine which column group this belongs to
                    # COL1 is columns 2-3, COL2 is columns 4-5, COL3 is columns 6-7, COL4 is columns 8-9
                    if col_idx in [2, 3]:  # COL1
                        cell.fill = col1_fill
                        cell.font = student_font
                        cell.alignment = center_align
                    elif col_idx in [4, 5]:  # COL2
                        cell.fill = col2_fill
                        cell.font = student_font
                        cell.alignment = center_align
                    elif col_idx in [6, 7]:  # COL3
                        cell.fill = col3_fill
                        cell.font = student_font
                        cell.alignment = center_align
                    elif col_idx in [8, 9]:  # COL4
                        cell.fill = col4_fill
                        cell.font = student_font
                        cell.alignment = center_align
            
            # Set column widths
            worksheet.column_dimensions[get_column_letter(1)].width = 20  # First column (labels)
            for col_idx in range(2, max_col + 1):
                col_letter = get_column_letter(col_idx)
                worksheet.column_dimensions[col_letter].width = 12  # Student roll number columns
            
            # Set row heights
            for row_idx in room_header_rows:
                worksheet.row_dimensions[row_idx].height = 30
            for row_idx in section_header_rows:
                worksheet.row_dimensions[row_idx].height = 22
            for row_idx in window_rows:
                worksheet.row_dimensions[row_idx].height = 25
            for row_idx in door_rows:
                worksheet.row_dimensions[row_idx].height = 22
            for row_idx in col_header_rows:
                worksheet.row_dimensions[row_idx].height = 18
            
            # Set row heights for data rows
            for row_idx in range(1, max_row + 1):
                if row_idx not in processed_rows:
                    worksheet.row_dimensions[row_idx].height = 18
            
        except Exception as e:
            print(f"    WARNING: Could not format seating sheet: {e}")
    
    def generate_seating_arrangements(self):
        """Generate seating arrangements for all exam classrooms."""
        print("\n" + "="*80)
        print("GENERATING SEATING ARRANGEMENTS")
        print("="*80)
        
        if not self.exam_classrooms:
            print("ERROR: No exam classrooms available")
            return False
        
        if self.student_data.empty:
            print("ERROR: No student data available")
            return False
        
        # Get all exam days and sessions
        exam_days = ['Saturday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Monday']
        sessions = ['FN', 'AN']
        
        filename = "seating arrangement.xlsx"
        filepath = os.path.join(OUTPUT_DIR, filename)
        
        # Create output directory if needed
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        
        try:
            writer = pd.ExcelWriter(filepath, engine='openpyxl')
            
            sheets_created = 0
            
            with writer as w:
                # Pre-compute students for each day/session (cache to avoid repeated lookups)
                students_cache = {}
                for day in exam_days:
                    for session in sessions:
                        students = self._get_students_for_exam(day, session)
                        students_cache[(day, session)] = students
                
                print(f"  Generating one sheet per room (total: {len(self.exam_classrooms)} rooms)")
                
                for room_name in self.exam_classrooms:
                    capacity = self.classroom_capacities.get(room_name, 50)
                    
                    # Collect seating data for all day/session combinations for this room
                    seating_data_by_day_session = {}
                    
                    for day in exam_days:
                        for session in sessions:
                            # Use cached students list
                            students = students_cache[(day, session)]
                            # Generate seating for this room, day, session
                            seating_df = self._generate_seating_for_room_with_students(room_name, capacity, students)
                            if not seating_df.empty:
                                seating_data_by_day_session[(day, session)] = seating_df
                    
                    # Create one combined sheet for this room
                    if seating_data_by_day_session:
                        sheet_df = self._create_seating_sheet(room_name, seating_data_by_day_session, exam_days)
                        
                        # Sheet name is just the room name
                        sheet_name = room_name
                        # Excel sheet name limit is 31 characters
                        if len(sheet_name) > 31:
                            sheet_name = sheet_name[:31]
                        
                        sheet_df.to_excel(w, sheet_name=sheet_name, index=False, header=False)
                        
                        # Apply color coding and formatting
                        try:
                            ws = w.sheets[sheet_name]
                            self._format_seating_sheet(ws, room_name, None, None)
                        except Exception as e:
                            print(f"    WARNING: Could not format sheet {sheet_name}: {e}")
                        
                        sheets_created += 1
                        if sheets_created % 5 == 0:
                            print(f"  Created {sheets_created} sheets...")
            
            print(f"\nSUCCESS: Created seating arrangement file")
            print(f"  File: {filepath}")
            print(f"  Total sheets: {sheets_created}")
            print(f"  Exam classrooms: {len(self.exam_classrooms)}")
            
            return True
            
        except Exception as e:
            print(f"ERROR: Could not create seating arrangement file: {e}")
            import traceback
            traceback.print_exc()
            return False

