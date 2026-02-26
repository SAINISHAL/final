"""Excel export utilities."""
import os
import time
import pandas as pd
from file_manager import FileManager
from config import DEPARTMENTS, TARGET_SEMESTERS, PRE_MID, POST_MID, DAYS, TEACHING_SLOTS
from excel_loader import ExcelLoader
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

class ExcelExporter:
    """Handles exporting of timetables to Excel files."""
    
    def __init__(self, data_frames, schedule_generator):
        self.dfs = data_frames
        self.schedule_gen = schedule_generator
        # Expanded vibrant pastel palette (readable on black text)
        self._palette = [
            "FFCDD2","F8BBD0","E1BEE7","D1C4E9","C5CAE9","BBDEFB","B3E5FC","B2EBF2",
            "B2DFDB","C8E6C9","DCEDC8","F0F4C3","FFF9C4","FFECB3","FFE0B2","FFCCBC",
            "D7CCC8","CFD8DC",
            "F28B82","F7A1C4","B39DDB","9FA8DA","90CAF9","81D4FA","80DEEA","80CBC4",
            "A5D6A7","C5E1A5","E6EE9C","FFF59D","FFE082","FFCC80","FFAB91",
            "AED581","81C784","4DD0E1","4FC3F7","9575CD","F48FB1"
        ]
        # Deterministic color mapping per exported workbook
        self._course_color_map = {}
    
    def _course_from_cell(self, val: str) -> str:
        """Extract a course identifier from a cell value."""
        if val is None:
            return ""
        s = str(val).strip()
        if not s or s == "-" or s.upper() == "FREE" or s.upper().startswith("LUNCH"):
            return ""
        # Common patterns: "CS161", "CS161 (Lab)", "CS161-Lab", "CS161: L"
        # Take up to first space or '(' or ':' or '-'
        for sep in [" (", " -", ":", " "]:
            if sep in s:
                s = s.split(sep)[0]
                break
        return s.strip()
    
    def _color_for_course(self, course: str) -> str:
        """Pick a stable color for the course within the current export."""
        if not course:
            return None
        if course not in self._course_color_map:
            idx = len(self._course_color_map) % len(self._palette)
            self._course_color_map[course] = self._palette[idx]
        return self._course_color_map[course]
    
    def _format_worksheet(self, worksheet, has_index=True, start_row=1):
        """Format worksheet to ensure all text is clearly visible.
        - Auto-adjusts column widths
        - Enables text wrapping
        - Sets appropriate row heights
        - Formats headers (bold, center)
        - Sets alignment for data cells"""
        try:
            # Find the maximum column and row with data
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            if max_row == 0 or max_col == 0:
                return
            
            # Format header row
            header_font = Font(bold=True, size=11)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Identify columns that likely contain course names or long text
            course_name_headers = ['course name', 'name', 'course_name']
            faculty_headers = ['faculty', 'instructor']
            time_slot_headers = ['time', 'slot', '08:00', '09:00', '10:00', '11:00', '12:00', '13:00', 
                                 '14:00', '15:00', '16:00', '17:00', '18:00', '19:00']
            
            # Get header row to identify column types
            header_row = {}
            for col_idx in range(1, max_col + 1):
                cell = worksheet.cell(row=start_row, column=col_idx)
                if cell.value is not None:
                    header_value = str(cell.value).lower().strip()
                    header_row[col_idx] = header_value
            
            # First pass: calculate optimal column widths
            column_widths = {}
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                avg_length = 0
                count = 0
                
                # Check all cells in this column
                for row_idx in range(start_row, max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        cell_value = str(cell.value)
                        cell_length = len(cell_value)
                        max_length = max(max_length, cell_length)
                        avg_length += cell_length
                        count += 1
                
                if count > 0:
                    avg_length = avg_length / count
                
                # Determine if this is a course name, faculty, or time slot column
                is_course_name_col = False
                is_faculty_col = False
                is_time_slot_col = False
                header_val = header_row.get(col_idx, '').lower()
                
                for course_name_keyword in course_name_headers:
                    if course_name_keyword in header_val:
                        is_course_name_col = True
                        break
                
                for faculty_keyword in faculty_headers:
                    if faculty_keyword in header_val:
                        is_faculty_col = True
                        break
                
                for time_keyword in time_slot_headers:
                    if time_keyword in header_val or ':' in header_val:
                        is_time_slot_col = True
                        break
                
                # Calculate column width based on content type
                if is_time_slot_col:
                    # Time slots are typically short (e.g., "08:00-08:30")
                    column_widths[col_letter] = max(12, min(max_length * 1.1, 18))
                elif is_course_name_col:
                    # Course names - give very generous width to minimize wrapping
                    # Use much wider columns for better readability
                    if max_length > 50:
                        # Very long course names (e.g., "DESIGN&ANALYSIS OF COMPUTER NETWORKS")
                        # Set width based on content length but ensure minimum of 50 units
                        # Approximate: each character is ~0.8 units for wrapped text
                        column_widths[col_letter] = min(max(50, max_length * 0.75), 60)
                    elif max_length > 30:
                        # Long course names - ensure wide enough to minimize wrapping
                        column_widths[col_letter] = min(max(40, max_length * 0.9), 55)
                    else:
                        # Medium length course names
                        column_widths[col_letter] = min(max(30, max_length * 1.1), 45)
                elif is_faculty_col:
                    # Faculty/Instructor columns - give good width for names
                    if max_length > 40:
                        # Long faculty names with multiple instructors
                        column_widths[col_letter] = min(max(35, max_length * 0.7), 50)
                    elif max_length > 25:
                        # Medium length faculty names
                        column_widths[col_letter] = min(max(30, max_length * 0.9), 45)
                    else:
                        # Short faculty names
                        column_widths[col_letter] = min(max(25, max_length * 1.1), 35)
                elif max_length > 40:
                    # Long text columns (but not course names) - still give good width
                    if max_length > 60:
                        column_widths[col_letter] = min(max(35, max_length * 0.6), 50)
                    else:
                        column_widths[col_letter] = min(max(30, max_length * 0.7), 45)
                else:
                    # Normal text columns
                    if max_length > 30:
                        # Medium length text - use reasonable width for wrapping
                        column_widths[col_letter] = min(max(15, max_length * 0.5), 35)
                    else:
                        # Short text - adjust based on content (1 character ≈ 1.1 units)
                        column_widths[col_letter] = min(max(10, max_length * 1.1), 30)
            
            # Second pass: apply formatting and column widths
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    col_letter = get_column_letter(col_idx)
                    
                    # Format header row
                    if row_idx == start_row:
                        cell.font = header_font
                        cell.alignment = header_alignment
                    else:
                        # Format data cells - wrap text
                        # Left align for course names and faculty, center align for others
                        header_val = header_row.get(col_idx, '').lower()
                        is_course_name = any(keyword in header_val for keyword in course_name_headers)
                        is_faculty = any(keyword in header_val for keyword in faculty_headers)
                        
                        if is_course_name or is_faculty:
                            # Left align course names and faculty for better readability
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        else:
                            # Center align other data
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    # Apply column width
                    if col_letter in column_widths:
                        worksheet.column_dimensions[col_letter].width = column_widths[col_letter]
            
            # Set row heights for better visibility
            # Header row
            if start_row <= max_row:
                worksheet.row_dimensions[start_row].height = 30
            
            # Data rows - set appropriate height for wrapped text
            for row_idx in range(start_row + 1, max_row + 1):
                # Check if row has any content
                has_content = False
                max_lines = 1
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        has_content = True
                        cell_value = str(cell.value)
                        col_letter = get_column_letter(col_idx)
                        col_width = column_widths.get(col_letter, 12)
                        # Estimate lines needed: approximately 10-12 characters per unit of width
                        chars_per_line = max(col_width * 0.85, 8)  # More conservative estimate
                        lines = max(1, len(cell_value) / chars_per_line)
                        max_lines = max(max_lines, lines)
                
                if has_content:
                    # Set height based on estimated lines (18 units per line for better spacing, minimum 25)
                    # Allow rows to be taller for multi-line content
                    row_height = max(25, min(18 * max_lines + 5, 80))
                    worksheet.row_dimensions[row_idx].height = row_height
            
            # Ensure index column (first column) is wide enough if it exists
            if has_index and max_col > 0:
                worksheet.column_dimensions['A'].width = max(15, worksheet.column_dimensions['A'].width or 15)
                
        except Exception as e:
            print(f"    WARNING: Could not format worksheet: {e}")
    
    def _apply_color_coding(self, worksheet, schedule_df, start_row=1, start_col=1):
        """Apply background colors and professional styling to department timetables."""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Professional Palette for Headers
        header_fill = PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid") # Indigo 500
        header_font = Font(color="FFFFFF", bold=True)
        
        index_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid") # Grey 100
        index_font = Font(bold=True)
        
        thin = Side(style='thin', color='B0BEC5')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Build mapping
        course_to_color = {}
        for day in schedule_df.index:
            for slot in schedule_df.columns:
                val = schedule_df.loc[day, slot]
                course = self._course_from_cell(val)
                if course and course not in course_to_color:
                    course_to_color[course] = self._color_for_course(course)

        header_rows = 1
        index_cols = 1
        nrows = len(schedule_df.index)
        ncols = len(schedule_df.columns)
        
        # 1. Style Headers (Row 1)
        for c in range(ncols + 1):
            cell = worksheet.cell(row=start_row, column=start_col + c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # 2. Style Index (Column 1) and Data
        for r in range(nrows):
            # Index cell (Day)
            idx_cell = worksheet.cell(row=start_row + header_rows + r, column=start_col)
            idx_cell.fill = index_fill
            idx_cell.font = index_font
            idx_cell.alignment = Alignment(horizontal='center', vertical='center')
            idx_cell.border = border
            
            for c in range(ncols):
                cell = worksheet.cell(row=start_row + header_rows + r, column=start_col + index_cols + c)
                cell.border = border
                val = cell.value
                course = self._course_from_cell(val)
                if course and course in course_to_color:
                    color = course_to_color[course]
                    try:
                        cell.fill = PatternFill(fill_type="solid", fgColor=color)
                    except Exception:
                        pass
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _apply_classroom_allocation_color_coding(self, worksheet):
        """Apply color coding to a classroom allocation sheet: header rows, day column, and course-based cell colors."""
        try:
            header_fill = PatternFill(fill_type="solid", fgColor="BBDEFB")   # Light blue for section headers
            slot_header_fill = PatternFill(fill_type="solid", fgColor="E3F2FD")  # Lighter blue for time slot row
            day_fill = PatternFill(fill_type="solid", fgColor="F5F5F5")     # Light gray for day column
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            if max_row < 1 or max_col < 1:
                return
            for row_idx in range(1, max_row + 1):
                cell_a = worksheet.cell(row=row_idx, column=1)
                val_a = cell_a.value
                val_str = str(val_a).strip() if val_a else ""
                # Semester section header row (e.g. "Semester 3")
                if val_str.startswith("Semester"):
                    for col_idx in range(1, max_col + 1):
                        c = worksheet.cell(row=row_idx, column=col_idx)
                        c.fill = header_fill
                    continue
                # Time slot header row (row after Semester: empty A, then slot labels)
                if row_idx <= max_row and val_str == "":
                    cell_b = worksheet.cell(row=row_idx, column=2).value
                    if cell_b and "-" in str(cell_b) and ":" in str(cell_b):  # e.g. 07:30-08:00
                        for col_idx in range(1, max_col + 1):
                            c = worksheet.cell(row=row_idx, column=col_idx)
                            c.fill = slot_header_fill
                        continue
                # Day row (MON, TUE, ...): light gray day column, course colors for slot cells
                if val_str in DAYS:
                    cell_a.fill = day_fill
                    for col_idx in range(2, max_col + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        val = cell.value
                        if val and str(val).strip():
                            course = self._course_from_cell(str(val))
                            if course:
                                color = self._color_for_course(course)
                                if color:
                                    cell.fill = PatternFill(fill_type="solid", fgColor=color)
        except Exception:
            pass
    
    def _get_course_details_for_session(self, semester, department, session_type):
        """Get course details for a specific department and session.
        Validates that expected courses from division logic match what should be scheduled."""
        try:
            # Get all semester courses
            sem_courses_all = ExcelLoader.get_semester_courses(self.dfs, semester)
            if sem_courses_all.empty:
                return pd.DataFrame()
            
            # Parse LTPSC
            sem_courses_parsed = ExcelLoader.parse_ltpsc(sem_courses_all)
            if sem_courses_parsed.empty:
                return pd.DataFrame()
            
            # Filter for department
            if 'Department' in sem_courses_parsed.columns:
                dept_mask = sem_courses_parsed['Department'].astype(str).str.contains(f"^{department}$", na=False, regex=True)
                dept_courses = sem_courses_parsed[dept_mask].copy()
            else:
                dept_courses = sem_courses_parsed.copy()
            
            if dept_courses.empty:
                return pd.DataFrame()
            
            # Divide by session
            pre_mid_courses, post_mid_courses = ExcelLoader.divide_courses_by_session(dept_courses, department, all_sem_courses=sem_courses_parsed)
            
            # Select appropriate session
            if session_type == PRE_MID:
                session_courses = pre_mid_courses
            else:
                session_courses = post_mid_courses
            
            if session_courses.empty:
                print(f"    WARNING: No courses assigned to {department} {session_type} session")
                return pd.DataFrame()
            
            # Prepare summary data
            summary_columns = ['Course Code', 'Course Name', 'Instructor', 'LTPSC', 'Lectures_Per_Week', 'Tutorials_Per_Week', 'Labs_Per_Week', 'Room Allocated', 'Lab Room Allocated', 'Combined Class']
            available_cols = [col for col in summary_columns if col in session_courses.columns]
            
            summary_df = session_courses[available_cols].copy()
            # Ensure Combined Class column exists
            if 'Combined Class' not in summary_df.columns:
                summary_df['Combined Class'] = 'NO'
            # Ensure Room Allocated columns exist
            if 'Room Allocated' not in summary_df.columns:
                summary_df['Room Allocated'] = ''
            if 'Lab Room Allocated' not in summary_df.columns:
                summary_df['Lab Room Allocated'] = ''
            
            # Format counts as "allocated/required" (e.g., "2/3" means 2 allocated out of 3 required)
            if 'Course Code' in summary_df.columns:
                for idx, row in summary_df.iterrows():
                    course_code = str(row.get('Course Code', '')).strip()
                    if course_code and course_code != 'nan' and course_code:
                        # Get required (expected) counts from course data
                        # Use 0 as default if column doesn't exist or value is missing
                        required_lectures = 0
                        required_tutorials = 0
                        required_labs = 0
                        
                        if 'Lectures_Per_Week' in summary_df.columns:
                            required_lectures = pd.to_numeric(row.get('Lectures_Per_Week', 0), errors='coerce')
                            if pd.isna(required_lectures):
                                required_lectures = 0
                            required_lectures = int(required_lectures)
                        
                        if 'Tutorials_Per_Week' in summary_df.columns:
                            required_tutorials = pd.to_numeric(row.get('Tutorials_Per_Week', 0), errors='coerce')
                            if pd.isna(required_tutorials):
                                required_tutorials = 0
                            required_tutorials = int(required_tutorials)
                        
                        if 'Labs_Per_Week' in summary_df.columns:
                            required_labs = pd.to_numeric(row.get('Labs_Per_Week', 0), errors='coerce')
                            if pd.isna(required_labs):
                                required_labs = 0
                            required_labs = int(required_labs)
                        
                        # Get actual allocated counts from schedule generator
                        actual = self.schedule_gen.get_actual_allocations(semester, department, session_type, course_code)
                        actual_lectures = actual.get('lectures', 0)
                        actual_tutorials = actual.get('tutorials', 0)
                        actual_labs = actual.get('labs', 0)
                        
                        # Get combined class flag from actual allocations (read from input Excel file)
                        combined_used = actual.get('combined_class', False)
                        # Room from schedule_gen allocations
                        room_alloc = actual.get('room', '')
                        lab_room_alloc = actual.get('lab_room', '')
                        # Format as "allocated/required"
                        if 'Lectures_Per_Week' in summary_df.columns:
                            summary_df.at[idx, 'Lectures_Per_Week'] = f"{actual_lectures}/{required_lectures}"
                        if 'Tutorials_Per_Week' in summary_df.columns:
                            summary_df.at[idx, 'Tutorials_Per_Week'] = f"{actual_tutorials}/{required_tutorials}"
                        if 'Labs_Per_Week' in summary_df.columns:
                            summary_df.at[idx, 'Labs_Per_Week'] = f"{actual_labs}/{required_labs}"
                        # Mark combined usage
                        summary_df.at[idx, 'Combined Class'] = 'YES' if combined_used else 'NO'
                        if 'Room Allocated' in summary_df.columns:
                            summary_df.at[idx, 'Room Allocated'] = room_alloc
                        if 'Lab Room Allocated' in summary_df.columns:
                            summary_df.at[idx, 'Lab Room Allocated'] = lab_room_alloc
            
            # Validate: Check if any courses have zero LTPSC (should still be included but may not schedule)
            # Note: Now checking actual allocated values from the formatted strings
            if 'Course Code' in summary_df.columns and 'Lectures_Per_Week' in summary_df.columns:
                # Extract actual values from "allocated/required" format for validation
                actual_lectures_list = []
                actual_tutorials_list = []
                actual_labs_list = []
                
                for idx, row in summary_df.iterrows():
                    # Extract allocated value from "allocated/required" format
                    lec_str = str(row.get('Lectures_Per_Week', '0/0'))
                    tut_str = str(row.get('Tutorials_Per_Week', '0/0'))
                    lab_str = str(row.get('Labs_Per_Week', '0/0'))
                    
                    # Parse "allocated/required" format
                    try:
                        actual_lec = int(lec_str.split('/')[0]) if '/' in lec_str else 0
                        actual_tut = int(tut_str.split('/')[0]) if '/' in tut_str else 0
                        actual_lab = int(lab_str.split('/')[0]) if '/' in lab_str else 0
                    except:
                        actual_lec = 0
                        actual_tut = 0
                        actual_lab = 0
                    
                    actual_lectures_list.append(actual_lec)
                    actual_tutorials_list.append(actual_tut)
                    actual_labs_list.append(actual_lab)
                
                # Check for zero LTPSC
                zero_ltpsc_mask = (
                    (pd.Series(actual_lectures_list) == 0) &
                    (pd.Series(actual_tutorials_list) == 0) &
                    (pd.Series(actual_labs_list) == 0)
                )
                zero_ltpsc = summary_df[zero_ltpsc_mask]
                if not zero_ltpsc.empty:
                    zero_codes = zero_ltpsc['Course Code'].dropna().tolist()
                    print(f"    INFO: {len(zero_codes)} courses with 0-0-0 LTPSC in {department} {session_type}: {', '.join(zero_codes)}")
            
            # Rename columns for better display
            column_rename = {
                'Lectures_Per_Week': 'Lectures/Week',
                'Tutorials_Per_Week': 'Tutorials/Week',
                'Labs_Per_Week': 'Labs/Week',
                'Instructor': 'Faculty'
            }
            summary_df = summary_df.rename(columns=column_rename)
            
            return summary_df
            
        except Exception as e:
            print(f"    WARNING: Could not generate course details: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()
    
    def _add_allocation_summary(self, writer, semester):
        """Add Allocation_Summary sheet (first): for this semester, total courses and how many allocated properly."""
        total_courses = 0
        allocated_properly = 0
        # track courses that didn't meet their required counts
        unallocated_list = []  # tuples: (department, session, course_code, req_lec, act_lec, req_tut, act_tut, req_lab, act_lab)
        # ensure storage for later retrieval
        if not hasattr(self, '_last_allocation_summary'):
            self._last_allocation_summary = {}
        try:
            for department in DEPARTMENTS:
                for session_type in [PRE_MID, POST_MID]:
                    session_courses = self._get_session_courses_for_summary(semester, department, session_type)
                    if session_courses is None or session_courses.empty:
                        continue
                    for _, row in session_courses.iterrows():
                        course_code = str(row.get('Course Code', '')).strip()
                        if not course_code or course_code == 'nan':
                            continue
                        req_lec = int(pd.to_numeric(row.get('Lectures_Per_Week', 0), errors='coerce') or 0)
                        req_tut = int(pd.to_numeric(row.get('Tutorials_Per_Week', 0), errors='coerce') or 0)
                        req_lab = int(pd.to_numeric(row.get('Labs_Per_Week', 0), errors='coerce') or 0)
                        actual = self.schedule_gen.get_actual_allocations(semester, department, session_type, course_code)
                        act_lec = actual.get('lectures', 0)
                        act_tut = actual.get('tutorials', 0)
                        act_lab = actual.get('labs', 0)
                        total_courses += 1
                        if act_lec >= req_lec and act_tut >= req_tut and act_lab >= req_lab:
                            allocated_properly += 1
                        else:
                            unallocated_list.append((department, session_type, course_code,
                                                    req_lec, act_lec, req_tut, act_tut,
                                                    req_lab, act_lab))
        except Exception as e:
            print(f"WARNING: Error computing allocation summary: {e}")
        summary_data = [
            ['Semester', semester],
            ['Total course allocations', total_courses],
            ['Allocated properly (L/T/P met)', allocated_properly],
            ['Not fully allocated', total_courses - allocated_properly],
        ]
        if total_courses > 0:
            summary_data.append(['Allocation %', f"{100 * allocated_properly / total_courses:.1f}%"])
        df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
        df.to_excel(writer, sheet_name='Allocation_Summary', index=False)
        # if there were any courses not fully allocated, add a details sheet
        if unallocated_list:
            cols = ['Department', 'Session', 'Course Code',
                    'Req Lec', 'Act Lec', 'Req Tut', 'Act Tut', 'Req Lab', 'Act Lab']
            df_fail = pd.DataFrame(unallocated_list, columns=cols)
            # sort by department/session for readability
            df_fail = df_fail.sort_values(['Department', 'Session', 'Course Code'])
            df_fail.to_excel(writer, sheet_name='Allocation_Failures', index=False)
            print(f"  NOTE: {len(unallocated_list)} courses were not fully allocated; see Allocation_Failures sheet.")
        try:
            ws = getattr(writer, 'book', None)
            if ws is not None and hasattr(ws, '__getitem__'):
                ws = ws['Allocation_Summary']
                self._format_worksheet(ws, has_index=False, start_row=1)
            elif hasattr(writer, 'sheets') and 'Allocation_Summary' in writer.sheets:
                self._format_worksheet(writer.sheets['Allocation_Summary'], has_index=False, start_row=1)
        except Exception:
            pass
        # save the results for caller to inspect
        try:
            self._last_allocation_summary[semester] = (total_courses, allocated_properly, unallocated_list)
        except Exception:
            pass
        print(f"  Allocation_Summary sheet created (Semester {semester}: {allocated_properly}/{total_courses} allocated properly)")

    def _get_session_courses_for_summary(self, semester, department, session_type):
        """Get session courses for a department/session (for allocation summary)."""
        try:
            sem_courses = ExcelLoader.get_semester_courses(self.dfs, semester)
            if sem_courses.empty:
                return None
            sem_courses = ExcelLoader.parse_ltpsc(sem_courses)
            if sem_courses.empty:
                return None
            if 'Department' in sem_courses.columns:
                dept_mask = sem_courses['Department'].astype(str).str.contains(f"^{department}$", na=False, regex=True)
                dept_courses = sem_courses[dept_mask].copy()
            else:
                dept_courses = sem_courses.copy()
            if dept_courses.empty:
                return None
            pre_mid, post_mid = ExcelLoader.divide_courses_by_session(dept_courses, department, all_sem_courses=sem_courses)
            session_courses = pre_mid if session_type == PRE_MID else post_mid
            return session_courses
        except Exception:
            return None

    def export_semester_timetable(self, semester):
        """Export timetable for a specific semester."""
        print(f"\n{'='*60}")
        print(f"GENERATING SEMESTER {semester} TIMETABLE")
        print(f"{'='*60}")
        self._course_color_map = {}
        # 1) Generate all schedules first so actual_allocations is populated for Allocation_Summary
        schedules = {}
        for department in DEPARTMENTS:
            print(f"\nProcessing {department}:")
            for session_type, label in [(PRE_MID, PRE_MID), (POST_MID, POST_MID)]:
                print(f"  {label} session...")
                try:
                    s = self.schedule_gen.generate_department_schedule(semester, department, session_type)
                    schedules[(department, session_type)] = s if s is not None else self.schedule_gen._initialize_schedule()
                except Exception as e:
                    print(f"    ERROR generating {department} {label}: {e}")
                    schedules[(department, session_type)] = self.schedule_gen._initialize_schedule()
        # 2) Open writer and write sheets
        filename = f"sem{semester}_timetable.xlsx"
        filepath = FileManager.get_output_path(filename)
        try:
            writer = pd.ExcelWriter(filepath, engine='openpyxl')
        except PermissionError:
            timestamp = int(time.time())
            alt_filename = f"sem{semester}_timetable_{timestamp}.xlsx"
            filepath = FileManager.get_output_path(alt_filename)
            writer = pd.ExcelWriter(filepath, engine='openpyxl')
            filename = alt_filename
        except Exception as e:
            print(f"\nFAILED: Could not create {filename}: {e}")
            import traceback
            traceback.print_exc()
            return False
        try:
            with writer as w:
                print(f"Creating {filename}...")
                try:
                    self._add_allocation_summary(w, semester)
                except Exception as e:
                    print(f"WARNING: Could not write Allocation_Summary: {e}")
                try:
                    self._add_course_summary(w, semester)
                except Exception as e:
                    print(f"WARNING: Could not write Course_Summary: {e}")
                department_count = 0
                for department in DEPARTMENTS:
                    for session_type, label in [(PRE_MID, PRE_MID), (POST_MID, POST_MID)]:
                        schedule = schedules.get((department, session_type))
                        if schedule is None:
                            continue
                        sheet_name = f"{department}_{label}"
                        clean_schedule = schedule.replace('Free', '-')
                        clean_schedule.to_excel(w, sheet_name=sheet_name, index=True, startrow=0)
                        try:
                            ws = w.sheets[sheet_name]
                            self._apply_color_coding(ws, clean_schedule, start_row=1, start_col=1)
                        except Exception as e:
                            print(f"    WARNING: Could not apply color coding to {sheet_name}: {e}")
                        course_details = self._get_course_details_for_session(semester, department, session_type)
                        if not course_details.empty:
                            start_row = len(clean_schedule) + 3
                            worksheet = w.sheets[sheet_name]
                            worksheet.cell(row=start_row, column=1, value="COURSE DETAILS:")
                            course_details.to_excel(w, sheet_name=sheet_name, index=False, startrow=start_row+1)
                        try:
                            ws = w.sheets[sheet_name]
                            self._format_worksheet(ws, has_index=True, start_row=1)
                        except Exception as e:
                            print(f"    WARNING: Could not format {sheet_name}: {e}")
                        department_count += 1
                        print(f"    SUCCESS: {sheet_name} created with course details")
                try:
                    self._add_electives_sheet(w, semester)
                except Exception as e:
                    print(f"WARNING: Could not add Electives sheet: {e}")
                try:
                    self._add_minor_sheet(w, semester)
                except Exception as e:
                    print(f"WARNING: Could not add Minor sheet: {e}")
                print(f"\nSUCCESS: Created {filename}")
                print(f"  - Allocation_Summary (first sheet)")
                print(f"  - {department_count} department schedules")
                print(f"  - Course summary sheet")
                print(f"  - Electives sheet")
                print(f"  - Minor sheet")
            return True
        except Exception as e:
            print(f"\nFAILED: Could not create {filename}: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _sanitize_sheet_name(self, name, max_len=31):
        """Excel sheet names: max 31 chars, no \\ / * ? : [ ]."""
        s = str(name).strip()
        for c in ['\\', '/', '*', '?', ':', '[', ']']:
            s = s.replace(c, '_')
        return s[:max_len] if len(s) > max_len else (s or 'Sheet')

    def _merge_combined_cell_entries(self, cells):
        """Merge same-course entries into one line, e.g. ['MA161 (DSAI)', 'MA161 (ECE)'] -> ['MA161 (DSAI, ECE)']."""
        from collections import defaultdict
        by_course = defaultdict(list)
        for c in cells:
            s = str(c).strip()
            if not s:
                continue
            if ' (' in s and s.endswith(')'):
                course = s.split(' (', 1)[0].strip()
                dept = s.split(' (', 1)[1][:-1].strip()
                by_course[course].append(dept)
            else:
                by_course[s].append('')
        out = []
        for course, depts in sorted(by_course.items()):
            depts_set = set(d for d in depts if d)
            
            # Special formatting for department groups
            if 'CSE-A' in depts_set and 'CSE-B' in depts_set:
                depts_set.remove('CSE-A')
                depts_set.remove('CSE-B')
                depts_set.add('CSE a and b')
            
            if 'DSAI' in depts_set and 'ECE' in depts_set:
                depts_set.remove('DSAI')
                depts_set.remove('ECE')
                depts_set.add('DSAI and ECE')
                
            depts_final = sorted(list(depts_set))
            if depts_final:
                out.append(f"{course} ({', '.join(depts_final)})")
            else:
                out.append(course)
        return out


    def _get_instructor_for_booking(self, semester_id, department, course_code, sem_courses=None):
        """Resolve instructor for a booking. For CSE-A/CSE-B, parses combined format when possible. Returns instructor name or empty string."""
        if sem_courses is None:
            sem_courses = ExcelLoader.get_semester_courses(self.dfs, semester_id)
        if sem_courses.empty:
            return ""
        dept_str = str(department).strip().upper()
        code_str = str(course_code).strip()
        if not code_str or code_str == "nan":
            return ""
        col_instructor = "Instructor" if "Instructor" in sem_courses.columns else ("Faculty" if "Faculty" in sem_courses.columns else None)
        if not col_instructor:
            return ""
        code_norm = code_str.upper()
        for _, row in sem_courses.iterrows():
            rdept = str(row.get("Department", "")).strip().upper()
            rcode = str(row.get("Course Code", "")).strip()
            rcode_norm = rcode.upper() if rcode else ""
            if rdept == dept_str and (rcode == code_str or rcode_norm == code_norm):
                val = row.get(col_instructor)
                if pd.notna(val) and str(val).strip():
                    val = str(val).strip()
                    # Try to parse A/B names regardless of current dept, 
                    # but only use the split name if the current dept matches A or B
                    name_a, name_b = ExcelLoader._parse_instructor_cse_ab(val)
                    
                    if dept_str == "CSE-A":
                        if name_a: return name_a
                        # If a label was found for B but not A, and this is A, 
                        # it might mean this teacher ONLY does B. 
                        # But wait, if name_b exists and name_a doesn't, we return the whole string 
                        # so that both get it? Or should we leave it?
                        # Standard: return whole string if no specific match
                        return val
                    
                    if dept_str == "CSE-B":
                        if name_b: return name_b
                        return val

                    # For combined CSE or any other department
                    return val
                return ""
        return ""

    def _faculty_sheet_key(self, instructor, department):
        """Use instructor name only (no section suffix) so each faculty has one sheet with all their classes (CSE-A and CSE-B together)."""
        return instructor

    def _build_faculty_sem_grid(self, room_bookings):
        """Build faculty_name -> (sem_id, session) -> (day, slot) -> list of cell text (e.g. 'CS262 (CSE-A)').
        Faculty key is instructor name only so one sheet per person (CSE-A and CSE-B classes on same sheet)."""
        from collections import defaultdict
        faculty_sem_grid = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
        # Cache semester courses per semester to avoid repeated get_semester_courses
        sem_courses_cache = {}
        for semester_key, semester_data in room_bookings.items():
            sem_id = semester_key.replace("sem_", "") if semester_key.startswith("sem_") else semester_key
            if sem_id not in sem_courses_cache:
                sem_courses_cache[sem_id] = ExcelLoader.get_semester_courses(self.dfs, int(sem_id) if str(sem_id).isdigit() else sem_id)
            sem_courses = sem_courses_cache[sem_id]
            for (day, slot), bookings in semester_data.items():
                for b in bookings:
                    session = b.get("session", "")
                    dept = b.get("dept", "")
                    course = str(b.get("course", "")).strip()
                    if not course or course == "nan":
                        continue
                    instructor = self._get_instructor_for_booking(sem_id, dept, course, sem_courses)
                    if not instructor:
                        instructor = "Unassigned"
                    faculty_key = self._faculty_sheet_key(instructor, dept)
                    cell_text = f"{course} ({dept})".strip()
                    key = (sem_id, session)
                    faculty_sem_grid[faculty_key][key][(day, slot)].append(cell_text)
        return dict(faculty_sem_grid)

    def _build_room_sem_grid_for_session(self, room_bookings, session_filter):
        """Build room -> semester -> (day, slot) -> list of cell text, including only bookings for session_filter (Pre-Mid or Post-Mid)."""
        from collections import defaultdict
        room_sem_grid = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
        for semester_key, semester_data in room_bookings.items():
            sem_id = semester_key.replace('sem_', '') if semester_key.startswith('sem_') else semester_key
            for (day, slot), bookings in semester_data.items():
                for b in bookings:
                    if b.get('session', '') != session_filter:
                        continue
                    room = b.get('room', '')
                    if not room:
                        continue
                    dept = b.get('dept', '')
                    course = str(b.get('course', '')).strip()
                    cell_text = f"{course} ({dept})".strip()
                    room_sem_grid[room][sem_id][(day, slot)].append(cell_text)
        return room_sem_grid

    def _classroom_allocation_summary_for_session(self, room_bookings, session_type):
        """Compute allocation summary for a session: total course allocations and how many have room assigned."""
        total = 0
        with_room = set()
        for semester in TARGET_SEMESTERS:
            for department in DEPARTMENTS:
                session_courses = self._get_session_courses_for_summary(semester, department, session_type)
                if session_courses is None or session_courses.empty:
                    continue
                for _, row in session_courses.iterrows():
                    course_code = str(row.get('Course Code', '')).strip()
                    if course_code and course_code != 'nan':
                        total += 1
        semester_key_prefix = 'sem_'
        for semester_key, semester_data in room_bookings.items():
            sem_id = semester_key.replace(semester_key_prefix, '') if semester_key.startswith(semester_key_prefix) else semester_key
            for (day, slot), bookings in semester_data.items():
                for b in bookings:
                    if b.get('session', '') != session_type:
                        continue
                    dept = b.get('dept', '')
                    course = str(b.get('course', '')).strip()
                    if dept and course:
                        with_room.add((sem_id, dept, course))
        return total, len(with_room)

    def _write_classroom_allocation_file(self, filename, all_rooms, room_sem_grid, session_label=None, allocation_summary=None):
        """Write one Excel file with Allocation_Summary (first) and one sheet per classroom from room_sem_grid.
        allocation_summary: (total, with_room) for session_label."""
        filepath = FileManager.get_output_path(filename)
        try:
            writer = pd.ExcelWriter(filepath, engine='openpyxl')
        except PermissionError:
            timestamp = int(time.time())
            alt_filename = f"{filename.replace('.xlsx', '')}_{timestamp}.xlsx"
            filepath = FileManager.get_output_path(alt_filename)
            writer = pd.ExcelWriter(filepath, engine='openpyxl')
            filename = alt_filename
        except Exception as e:
            print(f"FAILED: Could not create {filename}: {e}")
            return False
        try:
            with writer as w:
                if allocation_summary is not None and session_label is not None:
                    total, with_room = allocation_summary
                    not_allocated = total - with_room
                    summary_data = [
                        ['Session', session_label],
                        ['Total course allocations', total],
                        ['With room allocated', with_room],
                        ['Not allocated', not_allocated],
                    ]
                    if total > 0:
                        summary_data.append(['Allocation %', f"{100 * with_room / total:.1f}%"])
                    summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
                    summary_df.to_excel(w, sheet_name='Allocation_Summary', index=False)
                    try:
                        wb = getattr(w, 'book', None)
                        if wb is not None and hasattr(wb, '__getitem__'):
                            ws = wb['Allocation_Summary']
                            self._format_worksheet(ws, has_index=False, start_row=1)
                            # Light header color for Allocation_Summary
                            try:
                                for col in range(1, 3):
                                    ws.cell(row=1, column=col).fill = PatternFill(fill_type="solid", fgColor="E3F2FD")
                            except Exception:
                                pass
                    except Exception:
                        pass
                for room in all_rooms:
                    sheet_name = self._sanitize_sheet_name(room)
                    sem_data = room_sem_grid.get(room, {})
                    rows = []
                    for sem_id in sorted(sem_data.keys(), key=lambda x: (int(x) if str(x).isdigit() else 999, x)):
                        grid = sem_data[sem_id]
                        rows.append([f"Semester {sem_id}"])
                        rows.append([])
                        header = [""] + list(TEACHING_SLOTS)
                        rows.append(header)
                        for day in DAYS:
                            row = [day]
                            for slot in TEACHING_SLOTS:
                                cells = grid.get((day, slot), [])
                                merged = self._merge_combined_cell_entries(cells) if cells else []
                                cell_val = " | ".join(merged) if merged else ""
                                row.append(cell_val)
                            rows.append(row)
                        rows.append([])
                    if not rows:
                        rows = [["No allocations recorded for this room."]]
                    df = pd.DataFrame(rows)
                    df.to_excel(w, sheet_name=sheet_name, index=False, header=False)
                    try:
                        wb = getattr(w, 'book', None)
                        ws = wb[sheet_name] if wb is not None and hasattr(wb, '__getitem__') else getattr(w, 'sheets', {}).get(sheet_name)
                        if ws is not None:
                            self._format_worksheet(ws, has_index=False, start_row=1)
                            self._apply_classroom_allocation_color_coding(ws)
                    except Exception:
                        pass
                return True
        except Exception as e:
            print(f"FAILED: Could not write {filename}: {e}")
            import traceback
            traceback.print_exc()
            return False

    def export_classroom_view(self):
        """Create classroom_view.xlsx. One sheet per classroom, showing unified schedule with Pre/Post/Full indicators."""
        if not self.schedule_gen:
            return False
        room_bookings = getattr(self.schedule_gen, 'room_bookings', {})
        if not room_bookings:
            print("WARNING: No room bookings found; generate timetables first.")
            return False

        # Remove old files if they exist
        for f in ["classroom_allocation.xlsx", "classroom_allocation_pre_mid.xlsx", "classroom_allocation_post_mid.xlsx"]:
            old_path = FileManager.get_output_path(f)
            if os.path.exists(old_path):
                try:
                    os.remove(old_path)
                except Exception:
                    pass

        # 1. Aggregate bookings to determine Pre/Post/Full and Semester
        from collections import defaultdict
        cell_aggr = defaultdict(lambda: defaultdict(set))
        active_rooms = set()

        for sem_key, sem_data in room_bookings.items():
            # Extract semester id from sem_key "sem_X"
            semester_id = sem_key.replace("sem_", "")
            for (day, slot), bookings in sem_data.items():
                for b in bookings:
                    room = b.get('room')
                    if not room:
                        continue
                    active_rooms.add(room)
                    course = str(b.get('course', '')).strip()
                    dept = b.get('dept', '')
                    session = b.get('session', '') # 'Pre-Mid' or 'Post-Mid'
                    if course and dept:
                        # Store session AND semester
                        cell_aggr[(room, day, slot)][(course, dept)].add((session, semester_id))

        # 2. Build final grid data: room -> (day, slot) -> list of strings
        # String format: "Course [Session] [Sem X] (Dept)"
        room_grid = defaultdict(lambda: defaultdict(list))
        
        for (room, day, slot), course_map in cell_aggr.items():
            for (course, dept), sessions_sem_set in course_map.items():
                # sessions_sem_set has {(Pre-Mid, '3'), (Post-Mid, '3')} etc.
                sessions = {s[0] for s in sessions_sem_set}
                semesters = {s[1] for s in sessions_sem_set}
                sem_str = ",".join(sorted(semesters))
                
                label = ""
                if PRE_MID in sessions and POST_MID in sessions:
                    label = "Full"
                elif PRE_MID in sessions:
                    label = "Pre"
                elif POST_MID in sessions:
                    label = "Post"
                else:
                    label = "Unk"
                
                # Format: CS101 [Full] [Sem 3] (CSE-A)
                entry = f"{course} [{label}] [Sem {sem_str}] ({dept})"
                room_grid[room][(day, slot)].append(entry)


        # 3. Sort rooms
        # Provide a comprehensive list from schedule_gen, but filter by what we actually have or expected
        # Collect all definitions
        defined_rooms = []
        for name, _ in getattr(self.schedule_gen, 'classrooms', []):
            defined_rooms.append(name)
        c004 = getattr(self.schedule_gen, 'c004_room', None)
        if c004:
            defined_rooms.append(c004[0])
        
        # Union with active_rooms
        all_rooms = sorted(list(set(defined_rooms) | active_rooms), key=lambda x: (x.upper(), x))
        
        # 4. Write Excel
        filename = "classroom_view.xlsx"
        print(f"Creating {filename} with unified classroom schedules...")
        filepath = FileManager.get_output_path(filename)
        
        try:
            writer = pd.ExcelWriter(filepath, engine='openpyxl')
        except PermissionError:
            timestamp = int(time.time())
            filename = f"classroom_view_{timestamp}.xlsx"
            filepath = FileManager.get_output_path(filename)
            writer = pd.ExcelWriter(filepath, engine='openpyxl')
        except Exception as e:
            print(f"FAILED: Could not create {filename}: {e}")
            return False

        try:
            with writer as w:
                for room in all_rooms:
                    sheet_name = self._sanitize_sheet_name(room)
                    
                    # Prepare data for this room
                    rows = []
                    # Header
                    rows.append([f"Classroom: {room}"])
                    rows.append([])
                    header = [""] + list(TEACHING_SLOTS)
                    rows.append(header)
                    
                    grid = room_grid.get(room, {})
                    
                    for day in DAYS:
                        row = [day]
                        for slot in TEACHING_SLOTS:
                            cells = grid.get((day, slot), [])
                            # Use existing merge: "Course [Label] (Dept1, Dept2)"
                            # Logic in _merge_combined_cell_entries splits by ' ('
                            merged = self._merge_combined_cell_entries(cells) if cells else []
                            cell_val = " | ".join(merged) if merged else ""
                            row.append(cell_val)
                        rows.append(row)
                    
                    df = pd.DataFrame(rows)
                    df.to_excel(w, sheet_name=sheet_name, index=False, header=False)
                    
                    # Formatting
                    try:
                        wb = getattr(w, 'book', None)
                        ws = wb[sheet_name] if wb is not None and hasattr(wb, '__getitem__') else w.sheets[sheet_name]
                        if ws is not None:
                            self._format_worksheet(ws, has_index=False, start_row=1)
                            # Apply color coding? The generic one relies on course code matching raw text
                            # Our text is "Course [Label] (Dept)"
                            # _apply_classroom_allocation_color_coding expects text
                            self._apply_classroom_allocation_color_coding_unified(ws)
                    except Exception as e:
                        # print(f"    Warning formatting {sheet_name}: {e}")
                        pass
                
                print(f"SUCCESS: Created {filename} ({len(all_rooms)} sheets)")
                return True

        except Exception as e:
            print(f"FAILED: Could not write {filename}: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _apply_classroom_allocation_color_coding_unified(self, worksheet):
        """Apply premium color coding to unified views (Classroom/Faculty).
        Handles headers, day column, and course-based cell colors."""
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Professional Palette
        title_fill = PatternFill(start_color="455A64", end_color="455A64", fill_type="solid") # Blue Grey 700
        title_font = Font(color="FFFFFF", bold=True, size=12)
        
        header_fill = PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid") # Indigo 500
        header_font = Font(color="FFFFFF", bold=True)
        
        day_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid") # Grey 100
        day_font = Font(bold=True)
        
        thin = Side(style='thin', color='B0BEC5')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        max_row = worksheet.max_row
        max_col = worksheet.max_column
        if max_row < 3 or max_col < 1:
            return

        # 1. Title (Row 1)
        for c in range(1, max_col + 1):
            cell = worksheet.cell(row=1, column=c)
            cell.fill = title_fill
            cell.font = title_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # 2. Slot Headers (Row 3)
        for c in range(1, max_col + 1):
            cell = worksheet.cell(row=3, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

        # 3. Days (Column A) and Course Data
        for r in range(4, max_row + 1):
            # Day label
            day_cell = worksheet.cell(row=r, column=1)
            day_cell.fill = day_fill
            day_cell.font = day_font
            day_cell.alignment = Alignment(horizontal='center', vertical='center')
            day_cell.border = border
            
            # Data cells
            for c in range(2, max_col + 1):
                cell = worksheet.cell(row=r, column=c)
                cell.border = border
                val = cell.value
                if val and str(val).strip():
                    text = str(val).strip()
                    # Unified format: "Course [Label] ..." 
                    # Extract first part if multiple entries
                    first = text.split(" | ")[0]
                    # Extract course code: split by common separators
                    course_code = first
                    for sep in [" [", " (", " @", ":", " "]:
                        if sep in course_code:
                            course_code = course_code.split(sep)[0]
                    
                    course_code = course_code.strip()
                    color = self._color_for_course(course_code)
                    if color:
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

    def export_faculty_view(self):
        """Create faculty_view.xlsx. One sheet per faculty, showing unified schedule with Pre/Post/Full indicators."""
        if not self.schedule_gen:
            return False
        room_bookings = getattr(self.schedule_gen, 'room_bookings', {})
        if not room_bookings:
            print("WARNING: No room bookings found; generate timetables first.")
            return False

        # 1. Aggregate: faculty -> (day, slot) -> (course, dept, sem, room) -> set(sessions)
        from collections import defaultdict
        cell_aggr = defaultdict(lambda: defaultdict(lambda: defaultdict(set)))
        active_faculty = set()
        
        # Cache for instructors
        sem_courses_cache = {}
        
        for sem_key, sem_data in room_bookings.items():
            sem_id = sem_key.replace("sem_", "")
            if sem_id not in sem_courses_cache:
                 try:
                     sem_courses_cache[sem_id] = ExcelLoader.get_semester_courses(self.dfs, int(sem_id) if sem_id.isdigit() else sem_id)
                 except:
                     sem_courses_cache[sem_id] = pd.DataFrame()
            sem_courses = sem_courses_cache[sem_id]
            
            for (day, slot), bookings in sem_data.items():
                for b in bookings:
                    dept = b.get('dept', '')
                    course = str(b.get('course', '')).strip()
                    if not course or course == "nan":
                        continue
                    session = b.get('session', '')
                    room = b.get('room', '')
                    
                    instructors_str = self._get_instructor_for_booking(sem_id, dept, course, sem_courses)
                    if not instructors_str:
                        instructors_list = ["Unassigned"]
                    else:
                        # Split multiple instructors (comma, semicolon, ampersand, slash, pipe)
                        import re
                        instructors_list = [n.strip() for n in re.split(r'[,;&/|]', instructors_str) if n.strip()]
                        if not instructors_list:
                            instructors_list = ["Unassigned"]
                    
                    for faculty_key in instructors_list:
                        active_faculty.add(faculty_key)
                        cell_aggr[faculty_key][(day, slot)][(course, dept, sem_id, room)].add(session)

        # 2. Sort faculty
        all_faculty = sorted(list(active_faculty), key=lambda x: (x.upper(), x))
        
        # 3. Write Excel
        filename = "faculty_view.xlsx"
        print(f"Creating {filename} with unified faculty schedules...")
        filepath = FileManager.get_output_path(filename)
        
        try:
            writer = pd.ExcelWriter(filepath, engine='openpyxl')
        except PermissionError:
            timestamp = int(time.time())
            filename = f"faculty_view_{timestamp}.xlsx"
            filepath = FileManager.get_output_path(filename)
            writer = pd.ExcelWriter(filepath, engine='openpyxl')
        except Exception as e:
            print(f"FAILED: Could not create {filename}: {e}")
            return False

        try:
            with writer as w:
                for faculty in all_faculty:
                    sheet_name = self._sanitize_sheet_name(faculty)
                    
                    # Prepare data for this faculty
                    rows = []
                    # Header
                    rows.append([f"Faculty Timetable: {faculty}"])
                    rows.append([])
                    header = [""] + list(TEACHING_SLOTS)
                    rows.append(header)
                    
                    grid = cell_aggr.get(faculty, {})
                    
                    for day in DAYS:
                        row = [day]
                        for slot in TEACHING_SLOTS:
                            items = grid.get((day, slot), {})
                            cell_parts = []
                            # Sort by course code for consistency
                            for (course, dept, sem, room), sessions in sorted(items.items()):
                                label = ""
                                if PRE_MID in sessions and POST_MID in sessions:
                                    label = "Full"
                                elif PRE_MID in sessions:
                                    label = "Pre"
                                elif POST_MID in sessions:
                                    label = "Post"
                                else:
                                    label = "Unk"
                                
                                # Format: CS101 [Full] [Sem 1] (CSE-B) @ L106
                                cell_parts.append(f"{course} [{label}] [Sem {sem}] ({dept}) @ {room}")
                            
                            cell_val = " | ".join(cell_parts) if cell_parts else ""
                            row.append(cell_val)
                        rows.append(row)
                    
                    df = pd.DataFrame(rows)
                    df.to_excel(w, sheet_name=sheet_name, index=False, header=False)
                    
                    # Formatting
                    try:
                        wb = getattr(w, 'book', None)
                        ws = wb[sheet_name] if wb is not None and hasattr(wb, '__getitem__') else w.sheets[sheet_name]
                        if ws is not None:
                            self._format_worksheet(ws, has_index=False, start_row=1)
                            # Apply unified color coding (headers, days, course colors)
                            self._apply_classroom_allocation_color_coding_unified(ws)
                    except Exception:
                        pass
                
                print(f"SUCCESS: Created {filename} ({len(all_faculty)} faculty sheets)")
                return True

        except Exception as e:
            print(f"FAILED: Could not write {filename}: {e}")
            import traceback
            traceback.print_exc()
            return False


    def _add_course_summary(self, writer, semester):
        """Add course information summary. Always create the Course_Summary sheet (may be empty).
        Adds LTPSC validity check for all courses."""
        try:
            # Prepare empty default summary (columns if available)
            default_cols = ['Course Code', 'Course Name', 'LTPSC', 'Credits']
            summary_df = pd.DataFrame(columns=default_cols)

            ltpsc_valid_col = []
            all_valid = True

            if 'course' in self.dfs:
                course_df = self.dfs['course']
                if 'Semester' in course_df.columns:
                    temp_df = course_df.copy()
                    temp_df['Semester'] = pd.to_numeric(temp_df['Semester'], errors='coerce')
                    sem_courses = temp_df[temp_df['Semester'] == semester]

                    if not sem_courses.empty:
                        available_cols = [col for col in default_cols if col in sem_courses.columns]
                        summary_df = sem_courses[available_cols].copy()
                        # Check LTPSC validity for each course
                        for idx, row in summary_df.iterrows():
                            ltpsc_val = str(row.get('LTPSC', '')).strip()
                            valid = False
                            if ltpsc_val and '-' in ltpsc_val:
                                parts = ltpsc_val.split('-')
                                if len(parts) >= 3:
                                    try:
                                        float(parts[0])
                                        float(parts[1])
                                        float(parts[2])
                                        valid = True
                                    except Exception:
                                        valid = False
                            ltpsc_valid_col.append(valid)
                            if not valid:
                                all_valid = False
                        summary_df['LTPSC_Valid'] = ltpsc_valid_col
                        print(f"SUCCESS: Added Course_Summary sheet with {len(summary_df)} courses")
                    else:
                        print(f"WARNING: No courses found for semester {semester}; writing empty Course_Summary")
                else:
                    print("WARNING: 'Semester' column not found in course data; writing empty Course_Summary")
            else:
                print("WARNING: 'course' data frame not found; writing empty Course_Summary")

            # Add a message row at the top
            from pandas import DataFrame
            msg = "All courses follow LTPSC structure." if all_valid and not summary_df.empty else "Some courses do NOT follow LTPSC structure."
            msg_df = DataFrame({'Course Code': [msg]})
            # Write message row, then summary directly to the existing writer
            msg_df.to_excel(writer, sheet_name='Course_Summary', index=False, header=False, startrow=0)
            summary_df.to_excel(writer, sheet_name='Course_Summary', index=False, startrow=2)
            
            # Format Course_Summary worksheet
            try:
                ws = writer.sheets['Course_Summary']
                # Format the message row (row 1) separately
                if ws.max_row > 0:
                    msg_cell = ws.cell(row=1, column=1)
                    msg_cell.font = Font(bold=True, size=11)
                    msg_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    # Merge cells for message if there are multiple columns
                    if ws.max_column > 1:
                        try:
                            ws.merge_cells(f'A1:{get_column_letter(ws.max_column)}1')
                        except:
                            pass
                
                # Format the rest of the sheet (header at row 2, data starts at row 3)
                self._format_worksheet(ws, has_index=False, start_row=2)  # Header is at row 2
            except Exception as e:
                print(f"WARNING: Could not format Course_Summary: {e}")
        except Exception as e:
            print(f"FAILED: Could not add course summary: {e}")
    
    def _assign_room_by_capacity(self, students, semester_id, assigned_rooms=None):
        """Assign rooms for electives/minors based on student strength with deterministic rules."""
        if not self.schedule_gen:
            return ""
        
        try:
            students = int(float(students)) if students else 0
        except (ValueError, TypeError):
            return ""
        
        if students <= 0:
            return ""
        
        if assigned_rooms is None:
            assigned_rooms = set()
        
        # Build room pool preserving the original input order
        nonlab_rooms = list(getattr(self.schedule_gen, 'nonlab_rooms', []))
        room_pool = []
        seen = set()
        for name, cap in nonlab_rooms:
            if not name or name in seen:
                continue
            seen.add(name)
            room_pool.append((name, cap or 0))
        
        c004_room = getattr(self.schedule_gen, 'c004_room', None)
        if c004_room:
            room_pool.append(c004_room)
        
        if not room_pool and getattr(self.schedule_gen, 'classrooms', None):
            # Fall back to all classrooms (excluding labs), still preserving input order
            seen = set()
            lab_names = {r[0] for r in getattr(self.schedule_gen, 'lab_rooms', [])}
            for name, cap in self.schedule_gen.classrooms:
                if (not name or name in seen or name in lab_names):
                    continue
                seen.add(name)
                room_pool.append((name, cap or 0))
        
        if not room_pool:
            return ""
        
        shared_overlap = set(getattr(self.schedule_gen, 'shared_rooms', set()) or [])
        
        def can_reuse(room_name):
            if not room_name:
                return False
            return room_name.upper() in shared_overlap
        
        def first_matching_room(min_capacity=0):
            """Pick the first room (by input order) that:
            - is not already assigned (unless it's a shared/overlap room),
            - has capacity >= min_capacity (if capacity is known)."""
            for name, cap in room_pool:
                if not name:
                    continue
                if name in assigned_rooms and not can_reuse(name):
                    continue
                # If capacity is zero/unknown, still allow it; otherwise check threshold
                if cap and min_capacity:
                    if cap < min_capacity:
                        continue
                return name
            return ""

        # Prefer C004 for very large cohorts if configured
        if students >= 120 and c004_room:
            return c004_room[0]

        # For all other cases, simply pick the first suitable room in the input order
        # using a reasonable minimum capacity threshold.
        # We don't sort by capacity; we just walk the list in Excel order.
        min_cap = max(students, 0)
        room = first_matching_room(min_cap if min_cap > 0 else 0)
        if room:
            return room

        # Fallback: allow any remaining room regardless of capacity
        room = first_matching_room(0)
        return room or ""
    
    def _get_electives_data(self, semester):
        """Get elective data for a specific semester from 'Elective Data' sheet.
        Returns DataFrame with columns: Course Code, Course Name, Faculty, Semester, Students, Classroom"""
        try:
            # Try to find elective data sheet
            elective_df = None
            sheet_keys = [k for k in self.dfs.keys() if 'elective' in k.lower()]
            
            if not sheet_keys:
                # Try loading from course_data.xlsx directly
                try:
                    from config import INPUT_DIR
                    course_file = os.path.join(INPUT_DIR, 'course_data.xlsx')
                    if os.path.exists(course_file):
                        xl_file = pd.ExcelFile(course_file)
                        for sheet_name in xl_file.sheet_names:
                            if 'elective' in sheet_name.lower():
                                elective_df = pd.read_excel(course_file, sheet_name=sheet_name)
                                break
                except Exception as e:
                    pass
            
            if not sheet_keys and elective_df is None:
                # Check if it's in the loaded data frames with different naming
                for key in self.dfs.keys():
                    if 'elective' in key.lower() or 'electives' in key.lower():
                        elective_df = self.dfs[key]
                        break
            
            if sheet_keys and elective_df is None:
                elective_df = self.dfs[sheet_keys[0]]
            
            if elective_df is None or elective_df.empty:
                return pd.DataFrame(columns=['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom'])
            
            # Normalize column names - handle uppercase and variations
            elective_df = elective_df.copy()
            column_map = {}
            for col in elective_df.columns:
                col_lower = str(col).strip().lower()
                if any(x in col_lower for x in ['course code', 'code']) and 'name' not in col_lower:
                    column_map[col] = 'Course Code'
                elif any(x in col_lower for x in ['course name', 'coursename']) and 'code' not in col_lower:
                    column_map[col] = 'Course Name'
                elif any(x in col_lower for x in ['faculty', 'instructor', 'teacher']):
                    column_map[col] = 'Faculty'
                elif any(x in col_lower for x in ['semester', 'sem']):
                    column_map[col] = 'Semester'
                elif any(x in col_lower for x in ['student', 'registered', 'enrollment', 'enrol']):
                    column_map[col] = 'Students'
            
            elective_df = elective_df.rename(columns=column_map)
            
            # Filter by semester
            if 'Semester' in elective_df.columns:
                elective_df['Semester'] = pd.to_numeric(elective_df['Semester'], errors='coerce')
                elective_df = elective_df[elective_df['Semester'] == semester].copy()
            
            if elective_df.empty:
                return pd.DataFrame(columns=['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom'])
            
            # Prepare output columns
            output_cols = ['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students']
            available_cols = [col for col in output_cols if col in elective_df.columns]
            result_df = elective_df[available_cols].copy()
            
            # Add Classroom column and assign rooms - ensure each course gets a different room
            assigned_rooms = set()
            if 'Students' in result_df.columns:
                def assign_unique_room(students_val):
                    room = self._assign_room_by_capacity(students_val, semester, assigned_rooms)
                    if room:
                        assigned_rooms.add(room)
                    return room
                
                result_df['Classroom'] = result_df['Students'].apply(assign_unique_room)
            else:
                result_df['Classroom'] = ""
            
            # Ensure all required columns exist
            for col in ['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom']:
                if col not in result_df.columns:
                    result_df[col] = ""
            
            # Reorder columns
            result_df = result_df[['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom']]
            
            return result_df
            
        except Exception as e:
            print(f"ERROR: Could not load elective data: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame(columns=['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom'])
    
    def _get_minor_data(self, semester):
        """Get minor data for a specific semester from 'Minor Data' sheet.
        Returns DataFrame with columns: Course Code, Course Name, Faculty, Semester, Students, Classroom"""
        try:
            # Try to find minor data sheet
            minor_df = None
            sheet_keys = [k for k in self.dfs.keys() if 'minor' in k.lower()]
            
            if not sheet_keys:
                # Try loading from course_data.xlsx directly
                try:
                    from config import INPUT_DIR
                    course_file = os.path.join(INPUT_DIR, 'course_data.xlsx')
                    if os.path.exists(course_file):
                        xl_file = pd.ExcelFile(course_file)
                        for sheet_name in xl_file.sheet_names:
                            if 'minor' in sheet_name.lower():
                                minor_df = pd.read_excel(course_file, sheet_name=sheet_name)
                                break
                except Exception as e:
                    pass
            
            if not sheet_keys and minor_df is None:
                # Check if it's in the loaded data frames with different naming
                for key in self.dfs.keys():
                    if 'minor' in key.lower() or 'minors' in key.lower():
                        minor_df = self.dfs[key]
                        break
            
            if sheet_keys and minor_df is None:
                minor_df = self.dfs[sheet_keys[0]]
            
            if minor_df is None or minor_df.empty:
                return pd.DataFrame(columns=['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom'])
            
            # Normalize column names - handle uppercase and variations
            # Minor data has different structure: 'MINOR COURSE ' instead of separate Code/Name
            minor_df = minor_df.copy()
            column_map = {}
            for col in minor_df.columns:
                col_lower = str(col).strip().lower()
                # Handle "MINOR COURSE " - use as both Course Code and Course Name
                if any(x in col_lower for x in ['minor course', 'course']) and 'semester' not in col_lower and 'student' not in col_lower:
                    column_map[col] = 'Course Name'  # Use as course name, we'll extract code if needed
                elif any(x in col_lower for x in ['faculty', 'instructor', 'teacher']):
                    column_map[col] = 'Faculty'
                elif any(x in col_lower for x in ['semester', 'sem']):
                    column_map[col] = 'Semester'
                elif any(x in col_lower for x in ['student', 'registered', 'enrollment', 'enrol']):
                    column_map[col] = 'Students'
            
            minor_df = minor_df.rename(columns=column_map)
            
            # Filter by semester
            if 'Semester' in minor_df.columns:
                minor_df['Semester'] = pd.to_numeric(minor_df['Semester'], errors='coerce')
                minor_df = minor_df[minor_df['Semester'] == semester].copy()
            
            if minor_df.empty:
                return pd.DataFrame(columns=['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom'])
            
            # Prepare output columns
            output_cols = ['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students']
            available_cols = [col for col in output_cols if col in minor_df.columns]
            result_df = minor_df[available_cols].copy()
            
            # For Minor data, if we have Course Name but not Course Code, use Course Name as Code
            # (Minor courses often don't have separate codes)
            if 'Course Name' in result_df.columns and 'Course Code' not in result_df.columns:
                result_df['Course Code'] = result_df['Course Name'].apply(lambda x: str(x).strip().upper()[:8] if x and str(x).strip() != '' and str(x).lower() != 'nan' else "")
            
            # If Course Code exists but Course Name doesn't, use Code as Name
            if 'Course Code' in result_df.columns and 'Course Name' not in result_df.columns:
                result_df['Course Name'] = result_df['Course Code']
            
            # Add Classroom column and assign rooms - ensure each course gets a different room
            assigned_rooms = set()
            if 'Students' in result_df.columns:
                def assign_unique_room(students_val):
                    room = self._assign_room_by_capacity(students_val, semester, assigned_rooms)
                    if room:
                        assigned_rooms.add(room)
                    return room
                
                result_df['Classroom'] = result_df['Students'].apply(assign_unique_room)
            else:
                result_df['Classroom'] = ""
            
            # Ensure all required columns exist
            for col in ['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom']:
                if col not in result_df.columns:
                    result_df[col] = ""
            
            # Reorder columns
            result_df = result_df[['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom']]
            
            return result_df
            
        except Exception as e:
            print(f"ERROR: Could not load minor data: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame(columns=['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom'])
    
    def _add_electives_sheet(self, writer, semester):
        """Add Electives sheet to the workbook."""
        try:
            electives_df = self._get_electives_data(semester)
            
            if electives_df.empty:
                # Create empty sheet with headers
                electives_df = pd.DataFrame(columns=['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom'])
                print(f"  No elective data found for semester {semester} - creating empty Electives sheet")
            
            # Write to Excel
            electives_df.to_excel(writer, sheet_name='Electives', index=False)
            
            # Format the sheet
            try:
                ws = writer.sheets['Electives']
                self._format_worksheet(ws, has_index=False, start_row=1)
                print(f"  SUCCESS: Added Electives sheet ({len(electives_df)} courses)")
            except Exception as e:
                print(f"  WARNING: Could not format Electives sheet: {e}")
                
        except Exception as e:
            print(f"  WARNING: Could not add Electives sheet: {e}")
    
    def _add_minor_sheet(self, writer, semester):
        """Add Minor sheet to the workbook."""
        try:
            minor_df = self._get_minor_data(semester)
            
            if minor_df.empty:
                # Create empty sheet with headers
                minor_df = pd.DataFrame(columns=['Course Code', 'Course Name', 'Faculty', 'Semester', 'Students', 'Classroom'])
                print(f"  No minor data found for semester {semester} - creating empty Minor sheet")
            
            # Write to Excel
            minor_df.to_excel(writer, sheet_name='Minor', index=False)
            
            # Format the sheet
            try:
                ws = writer.sheets['Minor']
                self._format_worksheet(ws, has_index=False, start_row=1)
                print(f"  SUCCESS: Added Minor sheet ({len(minor_df)} courses)")
            except Exception as e:
                print(f"  WARNING: Could not format Minor sheet: {e}")
                
        except Exception as e:
            print(f"  WARNING: Could not add Minor sheet: {e}")
    

    def export_semester7_timetable(self):
        """Export special unified timetable for 7th semester with baskets.
        Creates:
        1. Main timetable showing baskets (7B1, 7B2, 7B3, 7B4) - 9:00 AM to 5:30 PM only, 2 classes per basket
        2. Basket assignments sheet (which courses go to which baskets)"""
        semester = 7
        print(f"\n{'='*60}")
        print(f"GENERATING SEMESTER {semester} UNIFIED TIMETABLE (BASKETS)")
        print(f"{'='*60}")
        # Reset color map for each workbook
        self._course_color_map = {}
        
        filename = f"sem{semester}_timetable.xlsx"
        filepath = FileManager.get_output_path(filename)
        
        # Attempt to open writer
        try:
            writer = pd.ExcelWriter(filepath, engine='openpyxl')
        except PermissionError as pe:
            print(f"\nWARNING: Cannot write to {filepath} (Permission denied / file may be open).")
            import time
            timestamp = int(time.time())
            alt_filename = f"sem{semester}_timetable_{timestamp}.xlsx"
            alt_filepath = FileManager.get_output_path(alt_filename)
            print(f"Attempting alternative filename: {alt_filename}")
            try:
                writer = pd.ExcelWriter(alt_filepath, engine='openpyxl')
                filepath = alt_filepath
                filename = alt_filename
            except Exception as e:
                print(f"\nFAILED: Could not create {filename}: {e}")
                import traceback
                traceback.print_exc()
                return False
        except Exception as e:
            print(f"\nFAILED: Could not create {filename}: {e}")
            import traceback
            traceback.print_exc()
            return False
        
        try:
            with writer as w:
                print(f"Creating {filename}...")
                
                # Get 7th semester courses
                # Prefer '7th sem' sheet if available, otherwise filter from main course sheet
                sem7_courses = pd.DataFrame()
                
                # 1. Check specific 7th sem sheet first
                sem7_sheet_key = None
                for key in self.dfs.keys():
                    key_lower = key.lower()
                    if ('7th' in key_lower and 'sem' in key_lower) or key_lower == '7th_sem_':
                        sem7_sheet_key = key
                        break
                
                if sem7_sheet_key and sem7_sheet_key in self.dfs:
                    sem7_courses = self.dfs[sem7_sheet_key].copy()
                    print(f"    Found 7th semester sheet: {sem7_sheet_key} with {len(sem7_courses)} courses")
                    # Standardize columns? We assume they have Course Code, etc.
                else:
                     # 2. Fallback to main course data
                    if 'course' in self.dfs:
                         course_df = self.dfs['course']
                         if 'Semester' in course_df.columns:
                            temp_df = course_df.copy()
                            temp_df['Semester'] = pd.to_numeric(temp_df['Semester'], errors='coerce')
                            sem7_courses = temp_df[temp_df['Semester'] == semester].copy()
                
                if sem7_courses.empty:
                    print("ERROR: No Semester 7 courses found (checked '7th sem' sheet and main 'Course Data')")
                    return False

                # Ensure Course Code exists
                cmd_col = None
                for col in sem7_courses.columns:
                    if str(col).lower() == 'course code':
                        cmd_col = col
                        break
                if not cmd_col: 
                     # Try to find header by content?
                     print("ERROR: 'Course Code' column not found in data")
                     return False
                
                sem7_courses.rename(columns={cmd_col: 'Course Code'}, inplace=True)

                # Identify baskets (pattern: 7B1, 7B2, 7B3, 7B4, etc.)
                # In 7th sem sheet, 'Basket Code' might be a separate column
                basket_col = None
                for col in sem7_courses.columns:
                    if 'basket' in str(col).lower():
                        basket_col = col
                        break
                
                baskets_set = set()
                if basket_col:
                     baskets_set = set(sem7_courses[basket_col].dropna().astype(str).unique())
                     # Filter out empty or 'nan'
                     baskets_set = {b for b in baskets_set if b.lower() not in ('nan', '', 'none')}
                else:
                    # Fallback: check Course Code for 7Bxxx pattern if no Basket column
                    basket_mask = sem7_courses['Course Code'].astype(str).str.match(r'^7B\d+', na=False)
                    baskets_set = set(sem7_courses[basket_mask]['Course Code'].unique())

                baskets = sorted(list(baskets_set))
                print(f"Found {len(baskets)} baskets: {', '.join(baskets)}")
                
                if not baskets:
                    print("WARNING: No baskets found. Proceeding with empty timetable.")
                
                # 1. Generate unified timetable with baskets
                from config import DAYS, TEACHING_SLOTS, LECTURE_DURATION, LUNCH_SLOTS
                
                # Filter slots to only include 9:00 AM to 5:30 PM
                sem7_slots = [s for s in TEACHING_SLOTS if s >= '09:00-09:30' and s <= '17:00-17:30']
                
                schedule = pd.DataFrame(index=DAYS, columns=sem7_slots)
                for day in DAYS:
                    for slot in sem7_slots:
                        schedule.loc[day, slot] = 'Free'
                
                # Mark lunch slots
                for day in DAYS:
                    for lunch_slot in LUNCH_SLOTS:
                        if lunch_slot in schedule.columns:
                            schedule.loc[day, lunch_slot] = 'LUNCH BREAK'
                
                import random
                
                # Track assigned slots for each basket: basket_code -> list of (day, slot)
                basket_slots_map = {}
                
                # Schedule each basket with 2 lectures per week
                for basket_code in baskets:
                    scheduled = 0
                    attempts = 0
                    max_attempts = 200
                    assigned_for_basket = []
                    
                    while scheduled < 2 and attempts < max_attempts:
                        attempts += 1
                        day = random.choice(DAYS)
                        # Avoid lunch slots
                        available_start_slots = [s for s in sem7_slots if s not in LUNCH_SLOTS]
                        if not available_start_slots:
                            continue
                        
                        start_slot = random.choice(available_start_slots)
                        try:
                            start_idx = sem7_slots.index(start_slot)
                            end_idx = start_idx + LECTURE_DURATION
                            if end_idx > len(sem7_slots):
                                continue
                            slots = sem7_slots[start_idx:end_idx]
                            
                            # Check if slots are free in LOCAL schedule
                            if all(schedule.loc[day, s] == 'Free' for s in slots):
                                if any(s in LUNCH_SLOTS for s in slots):
                                    continue
                                
                                # Assign basket
                                for slot in slots:
                                    schedule.loc[day, slot] = basket_code
                                    assigned_for_basket.append((day, slot))
                                scheduled += 1
                        except (ValueError, IndexError):
                            continue
                    
                    basket_slots_map[basket_code] = assigned_for_basket
                    if scheduled < 2:
                        print(f"    WARNING: Could not fully schedule basket {basket_code} (scheduled {scheduled}/2)")

                # Write main timetable
                clean_schedule = schedule.replace('Free', '-')
                clean_schedule.to_excel(w, sheet_name='Timetable', index=True, startrow=0)
                
                # Apply formatting
                try:
                    ws = w.sheets['Timetable']
                    self._apply_color_coding(ws, clean_schedule, start_row=1, start_col=1)
                    self._format_worksheet(ws, has_index=True, start_row=1)
                except Exception as e:
                    print(f"    WARNING: Formatting failed: {e}")
                
                print(f"    SUCCESS: Main timetable created")
                
                # 2. Create basket assignments (Course <-> Room Mapping)
                basket_assignments = pd.DataFrame(columns=['Basket Code', 'Course Code', 'Course Name', 'Department', 'LTPSC', 'Credits', 'Instructor', 'Classroom Allocated'])
                
                # Get non-lab classrooms
                nonlab_rooms = []
                if self.schedule_gen and hasattr(self.schedule_gen, 'nonlab_rooms'):
                    nonlab_rooms = [(name, cap) for name, cap in self.schedule_gen.nonlab_rooms if name and name.upper() != 'C004']
                
                # Fallback if empty
                if not nonlab_rooms and self.schedule_gen:
                    # try fetching all classrooms not in lab list
                    all_rooms = getattr(self.schedule_gen, 'classrooms', [])
                    lab_room_names = {r[0] for r in getattr(self.schedule_gen, 'lab_rooms', [])}
                    nonlab_rooms = [(n, c) for n, c in all_rooms if n not in lab_room_names and n.upper() != 'C004']

                if not nonlab_rooms:
                    print("WARNING: No classrooms available for assignment!")
                
                course_room_map = {}
                basket_used_rooms = {} # Basket -> Set of rooms used
                global_room_index = 0
                
                # Build list of courses in each basket
                # Iterate rows in sem7_courses
                for _, row in sem7_courses.iterrows():
                    # Determine basket
                    b_code = ''
                    if basket_col:
                        b_code = str(row.get(basket_col, '')).strip()
                    if not b_code and 'Course Code' in row:
                        cc = str(row['Course Code']).strip()
                        if cc.startswith('7B'):
                            b_code = cc
                    
                    if not b_code or b_code.lower() in ('nan', '', 'none'):
                        continue
                        
                    # Get Course Data
                    c_code = str(row.get('Course Code', '')).strip()
                    c_name = str(row.get('Course Name', '')).strip()
                    instructor = str(row.get('Instructor', '')).strip()
                    if not instructor and 'Faculty' in row:
                        instructor = str(row.get('Faculty', '')).strip()
                    
                    # Allocate Room
                    # Try to find a room that is:
                    # 1. Not used by other courses in THIS basket (since they run parallel)
                    # 2. Not used GLOBALLY at the times this basket is scheduled
                    
                    course_key = (b_code, c_code)
                    allocated_room = ''
                    
                    # Identify time slots for this basket
                    slots_for_basket = basket_slots_map.get(b_code, []) # List of (day, slot)
                    
                    # Set of rooms used by THIS basket
                    rooms_in_basket = basket_used_rooms.get(b_code, set())
                    
                    found_room = False
                    
                    # Iterate rooms cyclically
                    start_idx = global_room_index
                    for i in range(len(nonlab_rooms)):
                        idx = (start_idx + i) % len(nonlab_rooms)
                        r_name, r_cap = nonlab_rooms[idx]
                        
                        # Check local constraint: Unique room per course in basket
                        if r_name in rooms_in_basket:
                            continue
                        
                        # Check GLOBAL constraint: Room must be free at ALL basket slots
                        is_global_free = True
                        if self.schedule_gen:
                            # We must check ALL slots for ALL basket sessions
                            # slots_for_basket contains pairs (day, slot)
                            for day, slot in slots_for_basket:
                                # Check room_occupancy
                                occ_key = (day, slot)
                                if occ_key in self.schedule_gen.room_occupancy:
                                    # Check if r_name is used in ANY session (since Sem7 is 'Full')
                                    for r_u_name, r_u_sess, r_u_code in self.schedule_gen.room_occupancy[occ_key]:
                                        if r_u_name == r_name:
                                            is_global_free = False
                                            break
                                    if not is_global_free:
                                        break
                        
                        if is_global_free:
                            allocated_room = r_name
                            found_room = True
                            global_room_index = (idx + 1) % len(nonlab_rooms)
                            break
                    
                    if found_room:
                        # Register usage
                        basket_used_rooms.setdefault(b_code, set()).add(allocated_room)
                        course_room_map[course_key] = allocated_room
                        
                        # IMPORTANT: Mark as busy in global schedule generator
                        if self.schedule_gen:
                            for day, slot in slots_for_basket:
                                # Update room_occupancy
                                occ_key = (day, slot)
                                if occ_key not in self.schedule_gen.room_occupancy:
                                    self.schedule_gen.room_occupancy[occ_key] = set()
                                # Mark as 'Full' session for Semester 7 with course code
                                self.schedule_gen.room_occupancy[occ_key].add((allocated_room, 'Full', c_code))
                                
                                # Log booking for conflict checking
                                # We treat this as 'Lecture'
                                self.schedule_gen._log_room_booking(semester, day, slot, allocated_room, 'Sem7', c_code, 'Full')

                    else:
                        print(f"    WARNING: No available room found for {c_code} in basket {b_code}")
                        allocated_room = "UNALLOCATED"
                    
                    # Add to dataframe
                    basket_assignments = pd.concat([
                        basket_assignments,
                        pd.DataFrame([{
                            'Basket Code': b_code,
                            'Course Code': c_code,
                            'Course Name': c_name,
                            'Department': str(row.get('Department', '')),
                            'LTPSC': str(row.get('LTPSC', '')),
                            'Credits': str(row.get('Credits', '')),
                            'Instructor': instructor,
                            'Classroom Allocated': allocated_room
                        }])
                    ], ignore_index=True)

                basket_assignments.to_excel(w, sheet_name='Basket_Assignments', index=False)
                
                try:
                    ws = w.sheets['Basket_Assignments']
                    self._format_worksheet(ws, has_index=False, start_row=1)
                except Exception as e:
                    print(f"    WARNING: Formatting failed: {e}")

                print(f"    SUCCESS: Basket assignments sheet created with {len(basket_assignments)} entries")
                print(f"    Note: All assigned rooms have been registered in the global schedule to prevent conflicts.")
                
                return True
            
        except Exception as e:
            print(f"\nFAILED: Could not create {filename}: {e}")
            import traceback
            traceback.print_exc()
            return False
