"""Exam timetable scheduler for generating exam schedules."""
import pandas as pd
import random
from file_manager import FileManager
from excel_loader import ExcelLoader
from config import TARGET_SEMESTERS, DEPARTMENTS, PRE_MID, POST_MID, DAYS

class ExamScheduler:
    """Handles scheduling of exams for Pre-Mid and Post-Mid courses."""
    
    def __init__(self, data_frames, schedule_generator):
        self.dfs = data_frames
        self.schedule_gen = schedule_generator
        self.exam_classrooms = self._get_exam_classrooms()
        self.faculty_list = self._get_faculty_list()
    
    def get_all_pre_mid_courses(self):
        """Get all Pre-Mid courses from all target semesters."""
        all_pre_mid = []
        
        for semester in TARGET_SEMESTERS:
            # Get all semester courses
            sem_courses_all = ExcelLoader.get_semester_courses(self.dfs, semester)
            if sem_courses_all.empty:
                continue
            
            # Parse LTPSC
            sem_courses_parsed = ExcelLoader.parse_ltpsc(sem_courses_all)
            if sem_courses_parsed.empty:
                continue
            
            # Get Pre-Mid courses for each department
            for department in DEPARTMENTS:
                # Filter for department
                if 'Department' in sem_courses_parsed.columns:
                    dept_mask = sem_courses_parsed['Department'].astype(str).str.contains(f"^{department}$", na=False, regex=True)
                    dept_courses = sem_courses_parsed[dept_mask].copy()
                else:
                    dept_courses = sem_courses_parsed.copy()
                
                if dept_courses.empty:
                    continue
                
                # Divide by session
                pre_mid_courses, post_mid_courses = ExcelLoader.divide_courses_by_session(
                    dept_courses, department, all_sem_courses=sem_courses_parsed
                )
                
                # Add Pre-Mid courses (with semester and department info)
                if not pre_mid_courses.empty:
                    pre_mid_courses = pre_mid_courses.copy()
                    pre_mid_courses['Semester'] = semester
                    pre_mid_courses['Dept_Session'] = f"{department}_{PRE_MID}"
                    all_pre_mid.append(pre_mid_courses)
        
        if not all_pre_mid:
            return pd.DataFrame()
        
        # Combine all Pre-Mid courses
        all_pre_mid_df = pd.concat(all_pre_mid, ignore_index=True)
        
        # Remove duplicates by Course Code (same course in multiple departments)
        if 'Course Code' in all_pre_mid_df.columns:
            all_pre_mid_df = all_pre_mid_df.drop_duplicates(subset=['Course Code'], keep='first').reset_index(drop=True)
        
        return all_pre_mid_df
    
    def get_all_post_mid_courses(self):
        """Get all Post-Mid courses from all target semesters."""
        all_post_mid = []
        
        for semester in TARGET_SEMESTERS:
            # Get all semester courses
            sem_courses_all = ExcelLoader.get_semester_courses(self.dfs, semester)
            if sem_courses_all.empty:
                continue
            
            # Parse LTPSC
            sem_courses_parsed = ExcelLoader.parse_ltpsc(sem_courses_all)
            if sem_courses_parsed.empty:
                continue
            
            # Get Post-Mid courses for each department
            for department in DEPARTMENTS:
                # Filter for department
                if 'Department' in sem_courses_parsed.columns:
                    dept_mask = sem_courses_parsed['Department'].astype(str).str.contains(f"^{department}$", na=False, regex=True)
                    dept_courses = sem_courses_parsed[dept_mask].copy()
                else:
                    dept_courses = sem_courses_parsed.copy()
                
                if dept_courses.empty:
                    continue
                
                # Divide by session
                pre_mid_courses, post_mid_courses = ExcelLoader.divide_courses_by_session(
                    dept_courses, department, all_sem_courses=sem_courses_parsed
                )
                
                # Add Post-Mid courses (with semester and department info)
                if not post_mid_courses.empty:
                    post_mid_courses = post_mid_courses.copy()
                    post_mid_courses['Semester'] = semester
                    post_mid_courses['Dept_Session'] = f"{department}_{POST_MID}"
                    all_post_mid.append(post_mid_courses)
        
        if not all_post_mid:
            return pd.DataFrame()
        
        # Combine all Post-Mid courses
        all_post_mid_df = pd.concat(all_post_mid, ignore_index=True)
        
        # Remove duplicates by Course Code (same course in multiple departments)
        if 'Course Code' in all_post_mid_df.columns:
            all_post_mid_df = all_post_mid_df.drop_duplicates(subset=['Course Code'], keep='first').reset_index(drop=True)
        
        return all_post_mid_df
    
    def _get_exam_classrooms(self):
        """Get list of exam classrooms (non-lab, non-auditorium normal classrooms)."""
        exam_rooms = []
        if not self.schedule_gen or not self.schedule_gen.classrooms:
            return exam_rooms
        
        # Get classroom data to check room types
        classroom_df = self.dfs.get('classroom')
        if classroom_df is None or classroom_df.empty:
            # Fallback: use nonlab_rooms from schedule_gen
            exam_rooms = [room[0] for room in self.schedule_gen.nonlab_rooms if room[0] != 'C004']
            return exam_rooms
        
        # Find room number and type columns
        room_col = None
        type_col = None
        for col in classroom_df.columns:
            col_lower = str(col).lower()
            if room_col is None and any(k in col_lower for k in ['room', 'number', 'name']):
                room_col = col
            if type_col is None and any(k in col_lower for k in ['type', 'category']):
                type_col = col
        
        if room_col is None:
            room_col = classroom_df.columns[0]
        
        # Filter for normal classrooms (not lab, not auditorium)
        for _, row in classroom_df.iterrows():
            room_name = str(row.get(room_col, '')).strip()
            if not room_name:
                continue
            
            room_type = ''
            if type_col:
                room_type = str(row.get(type_col, '')).strip().lower()
            
            # Exclude labs and auditoriums
            if 'lab' in room_type or 'auditorium' in room_type:
                continue
            if 'lab' in room_name.lower() or room_name.upper() == 'C004':
                continue
            
            # Include normal classrooms
            if 'classroom' in room_type or room_name.startswith('C'):
                exam_rooms.append(room_name)
        
        # Remove duplicates and sort
        exam_rooms = sorted(list(set(exam_rooms)))
        return exam_rooms
    
    def _get_faculty_list(self):
        """Get list of faculty names from faculty_availability.xlsx."""
        faculty_list = []
        faculty_df = self.dfs.get('facultyavailability')
        if faculty_df is None or faculty_df.empty:
            # Try alternative key names
            for key in self.dfs.keys():
                if 'faculty' in key.lower():
                    faculty_df = self.dfs[key]
                    break
        
        if faculty_df is None or faculty_df.empty:
            print("WARNING: No faculty availability data found")
            return faculty_list
        
        # Find faculty name column
        faculty_col = None
        for col in faculty_df.columns:
            col_lower = str(col).lower()
            if 'faculty' in col_lower or 'name' in col_lower or 'instructor' in col_lower:
                faculty_col = col
                break
        
        if faculty_col is None:
            faculty_col = faculty_df.columns[0]
        
        # Extract unique faculty names
        faculty_list = faculty_df[faculty_col].dropna().unique().tolist()
        faculty_list = [str(f).strip() for f in faculty_list if str(f).strip()]
        faculty_list = sorted(list(set(faculty_list)))
        
        print(f"Loaded {len(faculty_list)} faculty members for invigilation")
        return faculty_list
    
    def _generate_invigilation_data(self, exam_days, num_classrooms_per_session=10):
        """Generate invigilation assignments for exam days.
        Returns DataFrame with columns: Day, Session, Classroom, Invigilator 1, Invigilator 2"""
        if not self.exam_classrooms:
            print("WARNING: No exam classrooms available for invigilation")
            return pd.DataFrame()
        
        if not self.faculty_list:
            print("WARNING: No faculty available for invigilation")
            return pd.DataFrame()
        
        invigilation_data = []
        sessions = ['FN', 'AN']
        
        for day in exam_days:
            for session in sessions:
                # Randomly select classrooms for this day/session
                # Use a subset of available classrooms (not all may be needed)
                num_rooms = min(num_classrooms_per_session, len(self.exam_classrooms))
                selected_rooms = random.sample(self.exam_classrooms, num_rooms)
                
                # Assign 2 invigilators per classroom
                available_faculty = self.faculty_list.copy()
                random.shuffle(available_faculty)
                
                faculty_idx = 0
                for room in selected_rooms:
                    # Get 2 different invigilators
                    invigilator1 = available_faculty[faculty_idx % len(available_faculty)]
                    faculty_idx += 1
                    invigilator2 = available_faculty[faculty_idx % len(available_faculty)]
                    faculty_idx += 1
                    
                    # Ensure different invigilators
                    if invigilator1 == invigilator2 and len(available_faculty) > 1:
                        invigilator2 = available_faculty[(faculty_idx) % len(available_faculty)]
                        faculty_idx += 1
                    
                    invigilation_data.append({
                        'Day': day,
                        'Session': session,
                        'Classroom': room,
                        'Invigilator 1': invigilator1,
                        'Invigilator 2': invigilator2
                    })
        
        invigilation_df = pd.DataFrame(invigilation_data)
        return invigilation_df
    
    def schedule_exams(self, courses_df, num_days=7):
        """Schedule exams across specified number of days.
        Ensures each day has exactly one exam from each department.
        Creates format with days as columns and all courses distributed.
        Returns two DataFrames: one for FN session, one for AN session."""
        if courses_df.empty:
            return pd.DataFrame(), pd.DataFrame()
        
        # Get unique course codes
        if 'Course Code' not in courses_df.columns:
            return pd.DataFrame(), pd.DataFrame()
        
        # Group courses by department
        courses_by_dept = {}
        if 'Department' in courses_df.columns:
            for _, row in courses_df.iterrows():
                course_code = str(row.get('Course Code', '')).strip()
                dept = str(row.get('Department', '')).strip()
                if course_code and dept:
                    if dept not in courses_by_dept:
                        courses_by_dept[dept] = []
                    if course_code not in courses_by_dept[dept]:
                        courses_by_dept[dept].append(course_code)
        else:
            # If no department column, treat all as one group
            all_courses = courses_df['Course Code'].dropna().unique().tolist()
            courses_by_dept['ALL'] = [str(c).strip() for c in all_courses if str(c).strip()]
        
        # Shuffle courses within each department for random distribution
        for dept in courses_by_dept:
            random.shuffle(courses_by_dept[dept])
        
        # Use 7 days: Saturday, Monday, Tuesday, Wednesday, Thursday, Friday, Monday
        exam_days = ['Saturday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Monday']
        exam_days = exam_days[:num_days]
        
        # Initialize schedule for FN and AN sessions
        # Each session is a dictionary: {day: [course1, course2, ...]}
        fn_schedule = {day: [] for day in exam_days}
        an_schedule = {day: [] for day in exam_days}
        
        # Track which course index we're at for each department
        dept_indices = {dept: 0 for dept in courses_by_dept}
        
        # Schedule exams ensuring one course per department per day
        for day_idx, day in enumerate(exam_days):
            # For each day, assign one course from each department
            # Balance between FN and AN sessions
            dept_list = list(courses_by_dept.keys())
            random.shuffle(dept_list)  # Randomize department order
            
            # Split departments between FN and AN (roughly half each)
            num_depts = len(dept_list)
            fn_depts = dept_list[:num_depts // 2] if num_depts > 1 else dept_list
            an_depts = dept_list[num_depts // 2:] if num_depts > 1 else []
            
            # Assign to FN session
            for dept in fn_depts:
                if dept_indices[dept] >= len(courses_by_dept[dept]):
                    continue  # No more courses for this department
                
                course = courses_by_dept[dept][dept_indices[dept]]
                dept_indices[dept] += 1
                fn_schedule[day].append(course)
            
            # Assign to AN session
            for dept in an_depts:
                if dept_indices[dept] >= len(courses_by_dept[dept]):
                    continue  # No more courses for this department
                
                course = courses_by_dept[dept][dept_indices[dept]]
                dept_indices[dept] += 1
                an_schedule[day].append(course)
        
        # After ensuring one per department per day, distribute remaining courses
        # Calculate remaining courses per department
        remaining_by_dept = {}
        for dept in courses_by_dept:
            remaining = len(courses_by_dept[dept]) - dept_indices[dept]
            if remaining > 0:
                remaining_by_dept[dept] = courses_by_dept[dept][dept_indices[dept]:]
        
        # Distribute remaining courses across days and sessions
        if remaining_by_dept:
            all_remaining = []
            for dept, courses in remaining_by_dept.items():
                for course in courses:
                    all_remaining.append((dept, course))
            
            random.shuffle(all_remaining)
            
            # Distribute remaining courses evenly
            for idx, (dept, course) in enumerate(all_remaining):
                day = exam_days[idx % len(exam_days)]
                # Alternate between FN and AN
                if idx % 2 == 0:
                    fn_schedule[day].append(course)
                else:
                    an_schedule[day].append(course)
        
        # Create DataFrames with days as columns
        # FN Session DataFrame
        fn_data = {}
        for day in exam_days:
            courses = fn_schedule.get(day, [])
            if courses:
                fn_data[day] = [', '.join([str(c) for c in courses])]
            else:
                fn_data[day] = ['']
        
        fn_df = pd.DataFrame(fn_data)
        
        # AN Session DataFrame
        an_data = {}
        for day in exam_days:
            courses = an_schedule.get(day, [])
            if courses:
                an_data[day] = [', '.join([str(c) for c in courses])]
            else:
                an_data[day] = ['']
        
        an_df = pd.DataFrame(an_data)
        
        return fn_df, an_df
    
    def _format_worksheet(self, worksheet, has_index=False, start_row=1):
        """Format worksheet to ensure all text is clearly visible."""
        try:
            from openpyxl.styles import Alignment, Font
            from openpyxl.utils import get_column_letter
            
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            if max_row == 0 or max_col == 0:
                return
            
            # Format header row
            header_font = Font(bold=True, size=11)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Calculate column widths
            column_widths = {}
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                
                for row_idx in range(1, max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        cell_value = str(cell.value)
                        max_length = max(max_length, len(cell_value))
                
                # Set appropriate width
                if max_length > 30:
                    column_widths[col_letter] = min(max(20, max_length * 0.6), 50)
                else:
                    column_widths[col_letter] = min(max(15, max_length * 1.2), 35)
            
            # Apply formatting
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    col_letter = get_column_letter(col_idx)
                    
                    if row_idx == start_row:
                        cell.font = header_font
                        cell.alignment = header_alignment
                    else:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    
                    if col_letter in column_widths:
                        worksheet.column_dimensions[col_letter].width = column_widths[col_letter]
            
            # Set row heights
            if start_row <= max_row:
                worksheet.row_dimensions[start_row].height = 30
            
            for row_idx in range(start_row + 1, max_row + 1):
                worksheet.row_dimensions[row_idx].height = 25
                
        except Exception as e:
            print(f"    WARNING: Could not format worksheet: {e}")
    
    def _format_exam_worksheet(self, worksheet):
        """Format exam worksheet with FN and AN sections."""
        try:
            from openpyxl.styles import Alignment, Font
            from openpyxl.utils import get_column_letter
            
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            if max_row == 0 or max_col == 0:
                return
            
            # Format headers (FN and AN section headers)
            header_font = Font(bold=True, size=12)
            section_font = Font(bold=True, size=11)
            cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Calculate column widths
            column_widths = {}
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                
                for row_idx in range(1, max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        cell_value = str(cell.value)
                        max_length = max(max_length, len(cell_value))
                
                # First column (labels) - narrower
                if col_idx == 1:
                    column_widths[col_letter] = max(20, min(max_length * 1.1, 30))
                else:
                    # Day columns - wider for course codes
                    column_widths[col_letter] = max(15, min(max_length * 0.8, 40))
            
            # Apply formatting
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    col_letter = get_column_letter(col_idx)
                    
                    cell_value = str(cell.value) if cell.value is not None else ''
                    
                    # Format section headers (FN and AN rows)
                    if 'FN:' in cell_value or 'AN:' in cell_value:
                        cell.font = section_font
                        cell.alignment = left_alignment
                    # Format day/date headers
                    elif row_idx in [2, 3, 6, 7] and col_idx > 1:
                        cell.font = Font(bold=True, size=10)
                        cell.alignment = cell_alignment
                    # Format course code cells
                    elif row_idx in [4, 8] and col_idx > 1:
                        cell.alignment = cell_alignment
                        # Set appropriate height for wrapped text
                        if cell_value:
                            lines = len(cell_value.split(','))
                            worksheet.row_dimensions[row_idx].height = max(25, 20 * lines)
                    # Format first column (labels)
                    elif col_idx == 1:
                        if 'Course Code' in cell_value:
                            cell.font = Font(bold=True, size=10)
                        cell.alignment = left_alignment
                    else:
                        cell.alignment = cell_alignment
                    
                    # Apply column width
                    if col_letter in column_widths:
                        worksheet.column_dimensions[col_letter].width = column_widths[col_letter]
            
            # Set row heights
            worksheet.row_dimensions[1].height = 25  # FN header
            worksheet.row_dimensions[2].height = 25  # FN Day row
            worksheet.row_dimensions[3].height = 25  # FN Date row
            worksheet.row_dimensions[4].height = 30  # FN Course codes (adjustable)
            worksheet.row_dimensions[5].height = 10  # Separator
            worksheet.row_dimensions[6].height = 25  # AN header
            worksheet.row_dimensions[7].height = 25  # AN Day row
            worksheet.row_dimensions[8].height = 25  # AN Date row
            worksheet.row_dimensions[9].height = 30  # AN Course codes (adjustable)
                
        except Exception as e:
            print(f"    WARNING: Could not format exam worksheet: {e}")
    
    def _create_exam_sheet(self, fn_df, an_df, exam_days):
        """Create a formatted exam sheet with FN and AN sections."""
        from datetime import datetime, timedelta
        import calendar
        
        # Find next Saturday from today (or use a fixed start date)
        # For now, use a sample date range (Sept 20, 2025 onwards)
        start_date = datetime(2025, 9, 20)  # Saturday, Sept 20, 2025
        
        # Create dates for each day
        dates = []
        current_date = start_date
        day_names_short = ['Sat', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Mon']
        day_names_full = ['Saturday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Monday']
        
        day_to_date = {}
        current_date = start_date
        
        for i, day in enumerate(exam_days):
            if i == 0:
                # First day (Saturday)
                day_to_date[day] = start_date
                current_date = start_date
            else:
                # Calculate days to add based on previous day
                prev_day = exam_days[i-1]
                if prev_day == 'Saturday':
                    days_to_add = 2  # Skip Sunday
                elif prev_day == 'Friday':
                    days_to_add = 3  # Skip weekend
                elif prev_day == 'Monday' and i > 1:
                    # If previous Monday was not first, calculate normally
                    days_to_add = 1
                else:
                    days_to_add = 1
                
                current_date += timedelta(days=days_to_add)
                day_to_date[day] = current_date
        
        # Create combined sheet structure
        # Section 1: FN (10:00 AM - 11:30 AM)
        fn_rows = []
        
        # FN Header
        fn_header = ['FN: 10:00 AM to 11:30 AM']
        for day in exam_days:
            fn_header.append('')
        fn_rows.append(fn_header)
        
        # FN Day row
        fn_day_row = ['']
        for day in exam_days:
            fn_day_row.append(day)
        fn_rows.append(fn_day_row)
        
        # FN Date row
        fn_date_row = ['']
        for day in exam_days:
            date = day_to_date[day]
            fn_date_row.append(date.strftime('%d-%m-%Y'))
        fn_rows.append(fn_date_row)
        
        # FN Course codes row
        fn_course_row = ['Course Code']
        for day in exam_days:
            if day in fn_df.columns and not fn_df.empty:
                courses = fn_df[day].iloc[0] if len(fn_df[day]) > 0 else ''
                fn_course_row.append(courses if courses else '')
            else:
                fn_course_row.append('')
        fn_rows.append(fn_course_row)
        
        # Empty row separator
        fn_rows.append([''] * (len(exam_days) + 1))
        
        # Section 2: AN (03:00 PM - 04:30 PM)
        an_rows = []
        
        # AN Header
        an_header = ['AN: 03:00 PM to 04:30 PM']
        for day in exam_days:
            an_header.append('')
        an_rows.append(an_header)
        
        # AN Day row
        an_day_row = ['']
        for day in exam_days:
            an_day_row.append(day)
        an_rows.append(an_day_row)
        
        # AN Date row
        an_date_row = ['']
        for day in exam_days:
            date = day_to_date[day]
            an_date_row.append(date.strftime('%d-%m-%Y'))
        an_rows.append(an_date_row)
        
        # AN Course codes row
        an_course_row = ['Course Code']
        for day in exam_days:
            if day in an_df.columns and not an_df.empty:
                courses = an_df[day].iloc[0] if len(an_df[day]) > 0 else ''
                an_course_row.append(courses if courses else '')
            else:
                an_course_row.append('')
        an_rows.append(an_course_row)
        
        # Combine all rows
        all_rows = fn_rows + an_rows
        
        # Create DataFrame
        # Find max columns
        max_cols = max(len(row) for row in all_rows) if all_rows else len(exam_days) + 1
        
        # Pad rows to same length
        padded_rows = []
        for row in all_rows:
            padded_row = row + [''] * (max_cols - len(row))
            padded_rows.append(padded_row)
        
        # Create column names
        column_names = [''] + [f'Day_{i}' for i in range(1, max_cols)]
        
        df = pd.DataFrame(padded_rows, columns=column_names[:max_cols])
        return df
    
    def export_exam_timetable(self):
        """Export exam timetable to Excel file with FN and AN sections."""
        print("\n" + "="*80)
        print("GENERATING EXAM TIMETABLE")
        print("="*80)
        
        filename = "exam_timetable.xlsx"
        filepath = FileManager.get_output_path(filename)
        
        try:
            # Get all Pre-Mid and Post-Mid courses
            print("\nCollecting Pre-Mid courses...")
            pre_mid_courses = self.get_all_pre_mid_courses()
            print(f"Found {len(pre_mid_courses)} unique Pre-Mid courses")
            
            print("\nCollecting Post-Mid courses...")
            post_mid_courses = self.get_all_post_mid_courses()
            print(f"Found {len(post_mid_courses)} unique Post-Mid courses")
            
            # Schedule exams (returns FN and AN DataFrames)
            print("\nScheduling Mid-Semester exams (Pre-Mid courses)...")
            mid_fn_df, mid_an_df = self.schedule_exams(pre_mid_courses, num_days=7)
            exam_days = ['Saturday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Monday']
            
            print("\nScheduling End-Semester exams (Post-Mid courses)...")
            end_fn_df, end_an_df = self.schedule_exams(post_mid_courses, num_days=7)
            
            # Create Excel writer
            writer = pd.ExcelWriter(filepath, engine='openpyxl')
            
            try:
                with writer as w:
                    # Create Mid-Semester sheet
                    if not mid_fn_df.empty or not mid_an_df.empty:
                        mid_sem_sheet = self._create_exam_sheet(mid_fn_df, mid_an_df, exam_days)
                        mid_sem_sheet.to_excel(w, sheet_name='mid sem', index=False, header=False)
                        print(f"  Written Mid-Semester schedule with {len(pre_mid_courses)} courses")
                    else:
                        # Create empty sheet
                        empty_df = pd.DataFrame(columns=exam_days)
                        empty_df.to_excel(w, sheet_name='mid sem', index=False)
                        print("  WARNING: No Pre-Mid courses found - created empty Mid-Semester sheet")
                    
                    # Create End-Semester sheet
                    if not end_fn_df.empty or not end_an_df.empty:
                        end_sem_sheet = self._create_exam_sheet(end_fn_df, end_an_df, exam_days)
                        end_sem_sheet.to_excel(w, sheet_name='end sem', index=False, header=False)
                        print(f"  Written End-Semester schedule with {len(post_mid_courses)} courses")
                    else:
                        # Create empty sheet
                        empty_df = pd.DataFrame(columns=exam_days)
                        empty_df.to_excel(w, sheet_name='end sem', index=False)
                        print("  WARNING: No Post-Mid courses found - created empty End-Semester sheet")
                    
                    # Format worksheets
                    try:
                        # Format mid sem sheet
                        ws_mid = w.sheets['mid sem']
                        self._format_exam_worksheet(ws_mid)
                        
                        # Format end sem sheet
                        ws_end = w.sheets['end sem']
                        self._format_exam_worksheet(ws_end)
                        
                        print("  Applied formatting to worksheets")
                    except Exception as e:
                        print(f"  WARNING: Could not format worksheets: {e}")
                    
                    # Create Invigilation Data sheet
                    invigilation_df = pd.DataFrame()
                    try:
                        print("\nGenerating Invigilation Data...")
                        invigilation_df = self._generate_invigilation_data(exam_days, num_classrooms_per_session=15)
                        if not invigilation_df.empty:
                            invigilation_df.to_excel(w, sheet_name='Invigilation Data', index=False)
                            print(f"  Created Invigilation Data sheet with {len(invigilation_df)} assignments")
                            print(f"    - {len(self.exam_classrooms)} exam classrooms available")
                            print(f"    - {len(self.faculty_list)} faculty available")
                            
                            # Format invigilation sheet
                            try:
                                ws_invig = w.sheets['Invigilation Data']
                                self._format_worksheet(ws_invig, has_index=False, start_row=1)
                                print("  Applied formatting to Invigilation Data sheet")
                            except Exception as e:
                                print(f"  WARNING: Could not format Invigilation Data sheet: {e}")
                        else:
                            print("  WARNING: Could not generate invigilation data")
                    except Exception as e:
                        print(f"  WARNING: Could not create Invigilation Data sheet: {e}")
                        import traceback
                        traceback.print_exc()
                
                print(f"\nSUCCESS: Created {filename}")
                print(f"  - Mid-Semester sheet: {len(pre_mid_courses)} courses distributed")
                print(f"  - End-Semester sheet: {len(post_mid_courses)} courses distributed")
                if not invigilation_df.empty:
                    print(f"  - Invigilation Data sheet: {len(invigilation_df)} assignments")
                print(f"  - File saved in: {FileManager.OUTPUT_DIR}")
                
                return True
                
            except Exception as e:
                print(f"ERROR: Could not write to Excel file: {e}")
                import traceback
                traceback.print_exc()
                return False
                
        except Exception as e:
            print(f"ERROR: Could not create exam timetable: {e}")
            import traceback
            traceback.print_exc()
            return False

